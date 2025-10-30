# =====================================
# app.py — Gestionale Clienti SHT (VERSIONE 2025 OTTIMIZZATA)
# =====================================
from __future__ import annotations
import streamlit as st
import pandas as pd
import time
from datetime import datetime
from pathlib import Path
from fpdf import FPDF
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, JsCode
from docx import Document
from docx.shared import Pt
# =====================================
# CONNESSIONE DATABASE MYSQL
# =====================================
import mysql.connector

def get_connection():
    """Crea la connessione al database MySQL usando le credenziali in secrets.toml"""
    return mysql.connector.connect(
        host=st.secrets["mysql"]["host"],
        database=st.secrets["mysql"]["database"],
        user=st.secrets["mysql"]["user"],
        password=st.secrets["mysql"]["password"]
    )

def load_table(table_name: str) -> pd.DataFrame:
    """Legge una tabella MySQL e la restituisce come DataFrame"""
    conn = get_connection()
    df = pd.read_sql(f"SELECT * FROM {table_name}", conn)
    conn.close()
    return df

def save_table(df: pd.DataFrame, table_name: str):
    """Aggiorna una tabella MySQL in modo sicuro (senza cancellare tutto)."""
    conn = get_connection()
    cur = conn.cursor()

    # Costruisci la query dinamica per ogni riga
    cols = df.columns.tolist()
    placeholders = ",".join(["%s"] * len(cols))
    col_list = ",".join([f"`{c}`" for c in cols])

    for _, row in df.iterrows():
        data = tuple(row)
        sql = f"REPLACE INTO {table_name} ({col_list}) VALUES ({placeholders})"
        cur.execute(sql, data)

    conn.commit()
    conn.close()


# =====================================
# CONFIGURAZIONE STREAMLIT E STILE BASE
# =====================================
st.set_page_config(page_title="GESTIONALE CLIENTI – SHT", layout="wide")

st.markdown("""
<style>
.block-container {
    padding-left: 2rem;
    padding-right: 2rem;
    max-width: 100% !important;
}
section.main > div:first-child {
    margin-top: 0 !important;
    padding-top: 0 !important;
}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<script>
    window.addEventListener('load', function() {
        window.scrollTo(0, 0);
    });
</script>
""", unsafe_allow_html=True)

# =====================================
# COSTANTI GLOBALI (VERSIONE GITHUB + STREAMLIT CLOUD)
# =====================================

APP_TITLE = "GESTIONALE CLIENTI – SHT"
LOGO_URL = "https://www.shtsrl.com/template/images/logo.png"

# Percorsi base relativi (funziona su GitHub e Streamlit Cloud)
STORAGE_DIR = Path(__file__).parent / "storage"
STORAGE_DIR.mkdir(parents=True, exist_ok=True)

# File CSV principali
CLIENTI_CSV = STORAGE_DIR / "clienti.csv"
CONTRATTI_CSV = STORAGE_DIR / "contratti.csv"

# Cartella e file per Gabriele
GABRIELE_DIR = STORAGE_DIR / "gabriele"
GABRIELE_DIR.mkdir(parents=True, exist_ok=True)
GABRIELE_CLIENTI = GABRIELE_DIR / "clienti.csv"
GABRIELE_CONTRATTI = GABRIELE_DIR / "contratti.csv"

# Cartella preventivi
PREVENTIVI_DIR = STORAGE_DIR / "preventivi"
PREVENTIVI_DIR.mkdir(parents=True, exist_ok=True)

# Cartella template preventivi
TEMPLATES_DIR = Path(__file__).parent / "templates"
TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
TEMPLATE_OPTIONS = {
    "Offerta A4": "Offerta_A4.docx",
    "Offerta A3": "Offerta_A3.docx",
    "Centralino": "Offerta_Centralino.docx",
    "Varie": "Offerta_Varie.docx",
}

# Durate standard contratti
DURATE_MESI = ["12", "24", "36", "48", "60", "72"]

# =====================================
# COLONNE STANDARD CSV
# =====================================
CLIENTI_COLS = [
    "ClienteID", "RagioneSociale", "PersonaRiferimento", "Indirizzo", "Citta", "CAP",
    "Telefono", "Cell", "Email", "PartitaIVA", "IBAN", "SDI",
    "UltimoRecall", "ProssimoRecall", "UltimaVisita", "ProssimaVisita",
    "TMK", "NoteCliente"
]


CONTRATTI_COLS = [
    "ClienteID", "RagioneSociale", "NumeroContratto", "DataInizio", "DataFine", "Durata",
    "DescrizioneProdotto", "NOL_FIN", "NOL_INT", "TotRata",
    "CopieBN", "EccBN", "CopieCol", "EccCol", "Stato"
]
# =====================================
# FUNZIONI UTILITY
# =====================================
def fmt_date(d) -> str:
    """Ritorna una data in formato DD/MM/YYYY"""
    import datetime as dt
    if d in (None, "", "nan", "NaN"):
        return ""
    try:
        if isinstance(d, (dt.date, dt.datetime, pd.Timestamp)):
            return pd.to_datetime(d).strftime("%d/%m/%Y")
        parsed = pd.to_datetime(str(d), errors="coerce", dayfirst=True)
        return "" if pd.isna(parsed) else parsed.strftime("%d/%m/%Y")
    except Exception:
        return ""

def money(x):
    """Formatta numeri in valuta italiana"""
    try:
        v = float(pd.to_numeric(x, errors="coerce"))
        if pd.isna(v): return ""
        return f"{v:,.2f} €"
    except Exception:
        return ""

def safe_text(txt):
    """Rimuove caratteri non compatibili con PDF latin-1"""
    if pd.isna(txt) or txt is None: return ""
    s = str(txt)
    replacements = {"€": "EUR", "–": "-", "—": "-", "“": '"', "”": '"', "‘": "'", "’": "'"}
    for k, v in replacements.items():
        s = s.replace(k, v)
    return s.encode("latin-1", "replace").decode("latin-1")

def ensure_columns(df, cols):
    for c in cols:
        if c not in df.columns:
            df[c] = pd.NA
    return df[cols]
def fix_inverted_dates(series: pd.Series, col_name: str = "") -> pd.Series:
    """
    Corregge automaticamente le date invertite (MM/DD/YYYY → DD/MM/YYYY)
    e mostra un log nel frontend Streamlit.
    """
    fixed = []
    fixed_count = 0
    total = len(series)

    for val in series:
        if pd.isna(val) or str(val).strip() == "":
            fixed.append("")
            continue

        s = str(val).strip()
        parsed = None

        try:
            # 1️⃣ Tentativo in formato italiano
            d1 = pd.to_datetime(s, dayfirst=True, errors="coerce")
            # 2️⃣ Tentativo in formato americano
            d2 = pd.to_datetime(s, dayfirst=False, errors="coerce")

            # Se entrambe valide e diverse → probabile inversione
            if not pd.isna(d1) and not pd.isna(d2) and d1 != d2:
                if d1.day <= 12 and d2.day > 12:
                    parsed = d2
                    fixed_count += 1
                else:
                    parsed = d1
            elif not pd.isna(d1):
                parsed = d1
            elif not pd.isna(d2):
                parsed = d2
            else:
                parsed = None
        except Exception:
            parsed = None

        if parsed is not None:
            fixed.append(parsed.strftime("%d/%m/%Y"))
        else:
            fixed.append("")

    # Mostra log su Streamlit (solo se ha corretto qualcosa)
    if fixed_count > 0:
        st.info(f"🔄 {fixed_count}/{total} date corrette automaticamente nella colonna **{col_name}**.")

    return pd.Series(fixed)
# =====================================
# BACKUP AUTOMATICO SU BOX (con controllo modifiche)
# =====================================
import requests, hashlib

def file_checksum(path: Path) -> str:
    """Calcola un hash (SHA1) del file per verificare se è cambiato."""
    if not path.exists():
        return ""
    h = hashlib.sha1()
    with open(path, "rb") as f:
        while chunk := f.read(8192):
            h.update(chunk)
    return h.hexdigest()

def box_upload_if_changed(local_path: Path, remote_name: str | None = None):
    """
    Carica automaticamente un file su Box solo se è cambiato rispetto all'ultima versione.
    Salva l'hash nel session_state per evitare duplicazioni.
    """
    try:
        folder_id = st.secrets["box"].get("backup_folder_id", "0")
        access_token = st.secrets["box"]["developer_token"]
        if not remote_name:
            remote_name = local_path.name

        # Calcolo hash attuale
        current_hash = file_checksum(local_path)
        prev_hash = st.session_state.get(f"box_hash_{local_path}", "")

        # Se non è cambiato → niente upload
        if current_hash == prev_hash:
            return

        url = "https://upload.box.com/api/2.0/files/content"
        headers = {"Authorization": f"Bearer {access_token}"}
        files = {
            "attributes": (None, f'{{"name":"{remote_name}","parent":{{"id":"{folder_id}"}}}}', "application/json"),
            "file": (remote_name, open(local_path, "rb")),
        }
        r = requests.post(url, headers=headers, files=files)
        if r.status_code in (200, 201):
            st.toast(f"☁️ Backup su Box completato: {remote_name}", icon="✅")
            st.session_state[f"box_hash_{local_path}"] = current_hash
        else:
            st.error(f"⚠️ Errore upload Box: {r.text}")
    except Exception as e:
        st.warning(f"⚠️ Backup Box non riuscito: {e}")

# =====================================
# CARICAMENTO E SALVATAGGIO DATI
# =====================================
def load_csv(path: Path, cols: list[str]) -> pd.DataFrame:
    if path.exists():
        df = pd.read_csv(path, dtype=str, encoding="utf-8-sig").fillna("")
    else:
        df = pd.DataFrame(columns=cols)
        df.to_csv(path, index=False, encoding="utf-8-sig")
    df = ensure_columns(df, cols)
    return df

def save_csv(df: pd.DataFrame, path: Path, date_cols=None):
    out = df.copy()
    if date_cols:
        for c in date_cols:
            out[c] = out[c].apply(fmt_date)
    out.to_csv(path, index=False, encoding="utf-8-sig")


def save_if_changed(df_new, path: Path, original_df):
    """Salva solo se i dati sono effettivamente cambiati."""
    import pandas as pd
    try:
        if not original_df.equals(df_new):
            df_new.to_csv(path, index=False, encoding='utf-8-sig')
            return True
        return False
    except Exception:
        df_new.to_csv(path, index=False, encoding='utf-8-sig')
        return True

# =====================================
# FUNZIONI DI SALVATAGGIO DEDICATE (con correzione automatica date)
# =====================================
def save_clienti(df: pd.DataFrame):
    """Salva il CSV clienti + backup su Box solo se modificato."""
    for c in ["UltimoRecall", "ProssimoRecall", "UltimaVisita", "ProssimaVisita"]:
        if c in df.columns:
            df[c] = fix_inverted_dates(df[c], col_name=c)
    save_csv(df, CLIENTI_CSV, date_cols=["UltimoRecall", "ProssimoRecall", "UltimaVisita", "ProssimaVisita"])
    box_upload_if_changed(CLIENTI_CSV)

def save_contratti(df: pd.DataFrame):
    """Salva il CSV contratti + backup su Box solo se modificato."""
    for c in ["DataInizio", "DataFine"]:
        if c in df.columns:
            df[c] = fix_inverted_dates(df[c], col_name=c)
    save_csv(df, CONTRATTI_CSV, date_cols=["DataInizio", "DataFine"])
    box_upload_if_changed(CONTRATTI_CSV)

# =====================================
# CONVERSIONE SICURA DATE ITALIANE (VERSIONE DEFINITIVA 2025)
# =====================================
from datetime import datetime

def parse_date_safe(val: str) -> str:
    """Converte una data in formato coerente DD/MM/YYYY, accettando formati misti."""
    if not val or str(val).strip() in ["nan", "NaN", "None", "NaT", ""]:
        return ""
    val = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(val, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return val


def to_date_series(series: pd.Series) -> pd.Series:
    """Compatibilità retroattiva: applica parse_date_safe a una serie pandas."""
    return series.apply(parse_date_safe)


# =====================================
# CARICAMENTO CLIENTI (senza salvataggio automatico)
# =====================================
def load_clienti() -> pd.DataFrame:
    """Carica i dati dei clienti dal file CSV (supporta ; , |)"""
    import pandas as pd
    if not CLIENTI_CSV.exists():
        return pd.DataFrame(columns=CLIENTI_COLS)

    for sep_try in [";", ",", "|", "\t"]:
        try:
            df = pd.read_csv(
                CLIENTI_CSV,
                dtype=str,
                sep=sep_try,
                encoding="utf-8-sig",
                on_bad_lines="skip",
                engine="python"
            ).fillna("")
            if len(df.columns) > 3:
                break
        except Exception:
            continue

    df = ensure_columns(df, CLIENTI_COLS)
    return df


# =====================================
# FUNZIONI DI CARICAMENTO DATI (VERSIONE DEFINITIVA 2025)
# =====================================

def normalize_cliente_id(df: pd.DataFrame) -> pd.DataFrame:
    """Normalizza la colonna ClienteID rimuovendo zeri iniziali e spazi."""
    if "ClienteID" not in df.columns:
        return df
    df["ClienteID"] = (
        df["ClienteID"]
        .astype(str)
        .str.strip()
        .str.replace(r"^0+", "", regex=True)
        .replace({"": None})
    )
    return df


def load_clienti() -> pd.DataFrame:
    """Carica i dati dei clienti dal file CSV (solo lettura, nessuna riscrittura automatica)."""
    import pandas as pd

    try:
        if CLIENTI_CSV.exists():
            df = pd.read_csv(
                CLIENTI_CSV,
                dtype=str,
                sep=None,              # autodetect ; or ,
                engine="python",
                encoding="utf-8-sig",
                on_bad_lines="skip"
            )
        else:
            df = pd.DataFrame(columns=CLIENTI_COLS)
    except Exception as e:
        st.error(f"❌ Errore durante la lettura dei clienti: {e}")
        df = pd.DataFrame(columns=CLIENTI_COLS)

    # Pulizia e normalizzazione
    df = (
        df.replace(to_replace=r"^(nan|NaN|None|NULL|null|NaT)$", value="", regex=True)
        .fillna("")
    )
    df = ensure_columns(df, CLIENTI_COLS)
    df = normalize_cliente_id(df)

    # Conversione date coerente
    for c in ["UltimoRecall", "ProssimoRecall", "UltimaVisita", "ProssimaVisita"]:
        if c in df.columns:
            df[c] = to_date_series(df[c])

    # ✅ Ordine alfabetico per Ragione Sociale
    df = df.sort_values("RagioneSociale", ascending=True, na_position="last").reset_index(drop=True)
    return df


def load_contratti() -> pd.DataFrame:
    """Carica i dati dei contratti dal file CSV (solo lettura, nessuna riscrittura automatica)."""
    import pandas as pd

    try:
        if CONTRATTI_CSV.exists():
            df = pd.read_csv(
                CONTRATTI_CSV,
                dtype=str,
                sep=None,              # autodetect ; or ,
                engine="python",
                encoding="utf-8-sig",
                on_bad_lines="skip"
            )
        else:
            df = pd.DataFrame(columns=CONTRATTI_COLS)
    except Exception as e:
        st.error(f"❌ Errore durante la lettura dei contratti: {e}")
        df = pd.DataFrame(columns=CONTRATTI_COLS)

    # Pulizia e normalizzazione
    df = (
        df.replace(to_replace=r"^(nan|NaN|None|NULL|null|NaT)$", value="", regex=True)
        .fillna("")
    )
    df = ensure_columns(df, CONTRATTI_COLS)
    df = normalize_cliente_id(df)

    # Conversione date coerente
    for c in ["DataInizio", "DataFine"]:
        if c in df.columns:
            df[c] = to_date_series(df[c])

    return df


# =====================================
# LOGIN FULLSCREEN
# =====================================
def do_login_fullscreen():
    """Login elegante con sfondo fullscreen"""
    if st.session_state.get("logged_in"):
        return st.session_state["user"], st.session_state["role"]

    st.markdown("""
    <style>
    div[data-testid="stAppViewContainer"] {padding-top:0 !important;}
    .block-container {
        display:flex;flex-direction:column;justify-content:center;
        align-items:center;height:100vh;background-color:#f8fafc;
    }
    .login-card {
        background:#fff;border:1px solid #e5e7eb;border-radius:12px;
        box-shadow:0 4px 16px rgba(0,0,0,0.08);
        padding:2rem 2.5rem;width:360px;text-align:center;
    }
    .login-title {font-size:1.3rem;font-weight:600;color:#2563eb;margin:1rem 0 1.4rem;}
    .stButton>button {
        width:260px;font-size:0.9rem;background-color:#2563eb;color:white;
        border:none;border-radius:6px;padding:0.5rem 0;
    }
    </style>
    """, unsafe_allow_html=True)

    login_col1, login_col2, _ = st.columns([1, 2, 1])
    with login_col2:
        st.markdown("<div class='login-card'>", unsafe_allow_html=True)
        st.image(LOGO_URL, width=140)
        st.markdown("<div class='login-title'>Accedi al CRM-SHT</div>", unsafe_allow_html=True)
        username = st.text_input("Nome utente", key="login_user").strip().lower()
        password = st.text_input("Password", type="password", key="login_pass")
        login_btn = st.button("Entra")
        st.markdown("</div>", unsafe_allow_html=True)

    # 🔹 Carica credenziali compatibili con formato Streamlit Cloud
    try:
        users = st.secrets["auth"]["users"]
    except Exception:
        # compatibilità con sottosezioni [auth.users.nome]
        users = st.secrets["auth"]["users"].to_dict() if hasattr(st.secrets["auth"]["users"], "to_dict") else st.secrets["auth"]["users"]
        if not users:
            users = {}
        # 🔹 costruisci manualmente il dizionario
        for k in st.secrets["auth"]["users"]:
            users[k] = st.secrets["auth"]["users"][k]

    if login_btn or (username and password and not st.session_state.get("_login_checked")):
        st.session_state["_login_checked"] = True

        # 🔹 compatibile con [auth.users.nome]
        if "auth" in st.secrets and "users" in st.secrets["auth"]:
            users = st.secrets["auth"]["users"]
        else:
            users = st.secrets.get("auth.users", {})

        # 🔹 login check
        if username in users and users[username]["password"] == password:
            st.session_state.update({
                "user": username,
                "role": users[username].get("role", "viewer"),
                "logged_in": True
            })
            st.success(f"✅ Benvenuto {username}!")
            time.sleep(0.3)
            st.rerun()
        else:
            st.error("❌ Credenziali non valide.")
            st.session_state["_login_checked"] = False

    st.stop()

# =====================================
# KPI CARD (riutilizzata)
# =====================================
def kpi_card(label: str, value, icon: str, color: str) -> str:
    return f"""
    <div style="
        background-color:{color};
        padding:18px;
        border-radius:12px;
        text-align:center;
        color:white;">
        <div style="font-size:26px;">{icon}</div>
        <div style="font-size:22px;font-weight:700;">{value}</div>
        <div style="font-size:14px;">{label}</div>
    </div>
    """

# =====================================
# PAGINA DASHBOARD (CLASSICA con TMK e gestione Fabio/Gabriele)
# =====================================
def page_dashboard(df_cli: pd.DataFrame, df_ct: pd.DataFrame, role: str):
    st.image(LOGO_URL, width=120)
    st.markdown("<h2>📊 Dashboard Gestionale</h2>", unsafe_allow_html=True)
    st.divider()

    # === KPI principali ===
    stato = df_ct["Stato"].fillna("").astype(str).str.lower()
    total_clients = len(df_cli)
    active_contracts = int((stato != "chiuso").sum())
    closed_contracts = int((stato == "chiuso").sum())
    now = pd.Timestamp.now().normalize()

    df_ct["DataInizio"] = pd.to_datetime(df_ct["DataInizio"], errors="coerce", dayfirst=True)
    new_contracts = df_ct[
        (df_ct["DataInizio"].notna()) &
        (df_ct["DataInizio"] >= pd.Timestamp(year=now.year, month=1, day=1))
    ]

    c1, c2, c3, c4 = st.columns(4)
    c1.markdown(kpi_card("Clienti attivi", total_clients, "👥", "#1976D2"), unsafe_allow_html=True)
    c2.markdown(kpi_card("Contratti attivi", active_contracts, "📄", "#388E3C"), unsafe_allow_html=True)
    c3.markdown(kpi_card("Contratti chiusi", closed_contracts, "❌", "#D32F2F"), unsafe_allow_html=True)
    c4.markdown(kpi_card("Nuovi contratti anno", len(new_contracts), "⭐", "#FBC02D"), unsafe_allow_html=True)
    st.divider()

    # === CREAZIONE NUOVO CLIENTE + CONTRATTO ===
    with st.expander("➕ Crea Nuovo Cliente + Contratto"):
        with st.form("frm_new_cliente"):
            st.markdown("#### 📇 Dati Cliente")

            col1, col2 = st.columns(2)
            with col1:
                ragione = st.text_input("🏢 Ragione Sociale")
                persona = st.text_input("👤 Persona Riferimento")
                indirizzo = st.text_input("📍 Indirizzo")
                citta = st.text_input("🏙️ Città")
                cap = st.text_input("📮 CAP")
                telefono = st.text_input("📞 Telefono")
                cell = st.text_input("📱 Cellulare")
            with col2:
                email = st.text_input("✉️ Email")
                piva = st.text_input("💼 Partita IVA")
                iban = st.text_input("🏦 IBAN")
                sdi = st.text_input("📡 SDI")
                note = st.text_area("📝 Note Cliente", height=70)
                tmk = st.selectbox(
                    "👩‍💼 TMK di riferimento",
                    ["", "Giulia", "Antonella", "Annalisa", "Laura"],
                    index=0
                )

            # === SEZIONE CONTRATTO ===
            st.markdown("#### 📄 Primo Contratto del Cliente")
            colc1, colc2, colc3 = st.columns(3)
            num = colc1.text_input("📄 Numero Contratto")
            data_inizio = colc2.date_input("📅 Data Inizio", format="DD/MM/YYYY")
            durata = colc3.selectbox("📆 Durata (mesi)", DURATE_MESI, index=2)
            desc = st.text_area("🧾 Descrizione Prodotto", height=80)
            
            colp1, colp2, colp3 = st.columns(3)
            nf = colp1.text_input("🏦 NOL_FIN")
            ni = colp2.text_input("🏢 NOL_INT")
            tot = colp3.text_input("💰 Tot Rata")

            colx1, colx2, colx3, colx4 = st.columns(4)
            with colx1:
                copie_bn = st.text_input("📄 Copie incluse B/N", value="", key="copie_bn")
            with colx2:
                ecc_bn = st.text_input("💰 Costo extra B/N (€)", value="", key="ecc_bn")
            with colx3:
                copie_col = st.text_input("🖨️ Copie incluse Colore", value="", key="copie_col")
            with colx4:
                ecc_col = st.text_input("💰 Costo extra Colore (€)", value="", key="ecc_col")

            # === SALVATAGGIO COMPLETO ===
            if st.form_submit_button("💾 Crea Cliente e Contratto"):
                try:
                    new_id = str(len(df_cli) + 1)
                    data_fine = pd.to_datetime(data_inizio) + pd.DateOffset(months=int(durata))

                    nuovo_cliente = {
                        "ClienteID": new_id,
                        "RagioneSociale": ragione,
                        "PersonaRiferimento": persona,
                        "Indirizzo": indirizzo,
                        "Citta": citta,
                        "CAP": cap,
                        "Telefono": telefono,
                        "Cell": cell,
                        "Email": email,
                        "PartitaIVA": piva,
                        "IBAN": iban,
                        "SDI": sdi,
                        "UltimoRecall": "",
                        "ProssimoRecall": "",
                        "UltimaVisita": "",
                        "ProssimaVisita": "",
                        "TMK": tmk,
                        "NoteCliente": note
                    }

                    nuovo_contratto = {
                        "ClienteID": new_id,
                        "RagioneSociale": ragione,
                        "NumeroContratto": num,
                        "DataInizio": fmt_date(data_inizio),
                        "DataFine": fmt_date(data_fine),
                        "Durata": durata,
                        "DescrizioneProdotto": desc,
                        "NOL_FIN": nf,
                        "NOL_INT": ni,
                        "TotRata": tot,
                        "CopieBN": copie_bn,
                        "EccBN": ecc_bn,
                        "CopieCol": copie_col,
                        "EccCol": ecc_col,
                        "Stato": "aperto"
                    }

                    # --- Scelta automatica file in base all'utente ---
                    user = st.session_state.get("user", "").lower()
                    if user == "gabriele":
                        path_cli = GABRIELE_CLIENTI
                        path_ct = GABRIELE_CONTRATTI
                    else:
                        path_cli = CLIENTI_CSV
                        path_ct = CONTRATTI_CSV

                    # --- Aggiorna CSV ---
                    if path_cli.exists():
                        df_exist = pd.read_csv(path_cli, dtype=str, encoding="utf-8-sig", on_bad_lines="skip").fillna("")
                    else:
                        df_exist = pd.DataFrame(columns=df_cli.columns)
                    df_exist = pd.concat([df_exist, pd.DataFrame([nuovo_cliente])], ignore_index=True)
                    df_exist.to_csv(path_cli, index=False, encoding="utf-8-sig")

                    if path_ct.exists():
                        df_ct_exist = pd.read_csv(path_ct, dtype=str, encoding="utf-8-sig", on_bad_lines="skip").fillna("")
                    else:
                        df_ct_exist = pd.DataFrame(columns=df_ct.columns)
                    df_ct_exist = pd.concat([df_ct_exist, pd.DataFrame([nuovo_contratto])], ignore_index=True)
                    df_ct_exist.to_csv(path_ct, index=False, encoding="utf-8-sig")

                    st.success(f"✅ Cliente '{ragione}' creato e salvato correttamente ({user.upper()})")
                    st.session_state.update({
                        "selected_cliente": new_id,
                        "nav_target": "Contratti",
                        "_go_contratti_now": True
                    })
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Errore creazione cliente: {e}")

    # === CONTRATTI IN SCADENZA ENTRO 6 MESI ===
    st.divider()
    st.markdown("### ⚠️ Contratti in scadenza entro 6 mesi")

    oggi = pd.Timestamp.now().normalize()
    entro_6_mesi = oggi + pd.DateOffset(months=6)
    df_ct["DataFine"] = pd.to_datetime(df_ct["DataFine"], errors="coerce", dayfirst=True)

    scadenze = df_ct[
        (df_ct["DataFine"].notna()) &
        (df_ct["DataFine"] >= oggi) &
        (df_ct["DataFine"] <= entro_6_mesi) &
        (df_ct["Stato"].astype(str).str.lower() != "chiuso")
    ].copy()

    if not scadenze.empty and "RagioneSociale" not in scadenze.columns:
        scadenze = scadenze.merge(df_cli[["ClienteID", "RagioneSociale"]], on="ClienteID", how="left")

    if scadenze.empty:
        st.success("✅ Nessun contratto attivo in scadenza nei prossimi 6 mesi.")
    else:
        scadenze["DataFine"] = scadenze["DataFine"].apply(fmt_date)
        scadenze = scadenze.sort_values("DataFine")
        st.markdown(f"📅 **{len(scadenze)} contratti in scadenza entro 6 mesi:**")
        head_cols = st.columns([2, 1, 1, 1, 0.8])
        head_cols[0].markdown("**Cliente**")
        head_cols[1].markdown("**Contratto**")
        head_cols[2].markdown("**Scadenza**")
        head_cols[3].markdown("**Stato**")
        head_cols[4].markdown("**Azioni**")
        st.markdown("---")

        for i, r in scadenze.iterrows():
            bg_color = "#f8fbff" if i % 2 == 0 else "#ffffff"
            # normalizzo l'ID da inviare
            raw_id = str(r.get("ClienteID", "")).strip()
            cliente_id_norm = raw_id.lstrip("0").upper()
        
            cols = st.columns([2, 1, 1, 1, 0.8])
            with cols[0]:
                st.markdown(f"<div style='background:{bg_color};padding:6px'><b>{r.get('RagioneSociale','—')}</b></div>", unsafe_allow_html=True)
            with cols[1]:
                st.markdown(f"<div style='background:{bg_color};padding:6px'>{r.get('NumeroContratto','—')}</div>", unsafe_allow_html=True)
            with cols[2]:
                st.markdown(f"<div style='background:{bg_color};padding:6px'>{fmt_date(r.get('DataFine'))}</div>", unsafe_allow_html=True)
            with cols[3]:
                st.markdown(f"<div style='background:{bg_color};padding:6px'>{r.get('Stato','—')}</div>", unsafe_allow_html=True)
            with cols[4]:
                if st.button("📂 Apri", key=f"open_scad_{cliente_id_norm}_{i}", use_container_width=True):
                    if cliente_id_norm:
                        st.session_state["selected_cliente"] = cliente_id_norm
                        st.session_state.pop("nav_target", None)
                        st.session_state["nav_target"] = "Contratti"
                        st.session_state["_go_contratti_now"] = True
                        st.rerun()
                    else:
                        st.warning("⚠️ ID cliente non valido per questo contratto.")


    # === CONTRATTI RECENTI SENZA DATA FINE ===
    st.divider()
    st.markdown("### ⚠️ Contratti recenti senza data di fine")

    df_ct["DataInizio"] = pd.to_datetime(df_ct["DataInizio"], errors="coerce", dayfirst=True)
    df_ct["DataFine"] = pd.to_datetime(df_ct["DataFine"], errors="coerce", dayfirst=True)
    oggi = pd.Timestamp.now().normalize()

    contratti_senza_fine = df_ct[
        (df_ct["DataFine"].isna()) &
        (df_ct["DataInizio"].notna()) &
        (df_ct["DataInizio"] >= oggi)
    ].copy()

    if contratti_senza_fine.empty:
        st.success("✅ Tutti i contratti recenti hanno una data di fine.")
    else:
        st.warning(f"⚠️ {len(contratti_senza_fine)} contratti inseriti da oggi non hanno ancora una data di fine:")
        if "RagioneSociale" not in contratti_senza_fine.columns:
            contratti_senza_fine = contratti_senza_fine.merge(
                df_cli[["ClienteID", "RagioneSociale"]],
                on="ClienteID", how="left"
            )

        contratti_senza_fine["DataInizio"] = contratti_senza_fine["DataInizio"].apply(fmt_date)
        contratti_senza_fine = contratti_senza_fine.sort_values("DataInizio", ascending=False)

        for i, r in contratti_senza_fine.iterrows():
            # normalizzo l'ID da inviare
            raw_id = str(r.get("ClienteID", "")).strip()
            cliente_id_norm = raw_id.lstrip("0").upper()
        
            col1, col2, col3, col4, col5 = st.columns([2.5, 1, 1.2, 2.5, 0.8])
            with col1:
                st.markdown(f"**{r.get('RagioneSociale', '—')}**")
                st.markdown(r.get("NumeroContratto", "—"))
            with col3:
                st.markdown(r.get("DataInizio", "—"))
            with col4:
                desc = str(r.get("DescrizioneProdotto", "—"))
                if len(desc) > 60:
                    desc = desc[:60] + "…"
                st.markdown(desc)
            with col5:
                if st.button("📂 Apri", key=f"open_ndf_{cliente_id_norm}_{i}", use_container_width=True):
                    if cliente_id_norm:
                        st.session_state["selected_cliente"] = cliente_id_norm
                        st.session_state["nav_target"] = "Contratti"
                        st.session_state["_go_contratti_now"] = True
                        st.rerun()
                    else:
                        st.warning("⚠️ ID cliente non valido per questo contratto.")




# =====================================
# PAGINA CLIENTI (con anagrafica visibile + editor, note e recall vicini)
# =====================================
def page_clienti(df_cli: pd.DataFrame, df_ct: pd.DataFrame, role: str):
    st.subheader("📋 Gestione Clienti")

    # === PRE-SELEZIONE CLIENTE DA NAVIGAZIONE ===
    if "selected_cliente" in st.session_state:
        sel_id_nav = str(st.session_state.pop("selected_cliente"))
        cli_ids = df_cli["ClienteID"].astype(str)
        if sel_id_nav in set(cli_ids):
            row = df_cli.loc[cli_ids == sel_id_nav].iloc[0]
            st.session_state["cliente_selezionato"] = row["RagioneSociale"]

    # === RICERCA CLIENTE ===
    search_query = st.text_input("🔍 Cerca cliente per nome o ID", key="search_cli")
    if search_query:
        filtered = df_cli[
            df_cli["RagioneSociale"].str.contains(search_query, case=False, na=False)
            | df_cli["ClienteID"].astype(str).str.contains(search_query, na=False)
        ]
    else:
        filtered = df_cli.copy()

    if filtered.empty:
        st.warning("❌ Nessun cliente trovato.")
        return

    options = filtered["RagioneSociale"].tolist()
    selected_name = st.session_state.get("cliente_selezionato", options[0])
    sel_rag = st.selectbox(
        "Seleziona Cliente",
        options,
        index=options.index(selected_name) if selected_name in options else 0,
        key="sel_cliente_box"
    )

    cliente = filtered[filtered["RagioneSociale"] == sel_rag].iloc[0]
    sel_id = str(cliente["ClienteID"])

    # === HEADER + AZIONI ===
    col1, col2 = st.columns([4, 1])
    with col1:
        st.markdown(f"## 🏢 {cliente['RagioneSociale']}")
        st.caption(f"ID Cliente: {sel_id}")
    with col2:
        if st.button("📄 Vai ai Contratti", use_container_width=True, key=f"go_cont_{sel_id}"):
            st.session_state.update({"selected_cliente": sel_id, "nav_target": "Contratti", "_go_contratti_now": True})
            st.rerun()

        if st.button("✏️ Modifica Anagrafica", use_container_width=True, key=f"btn_edit_{sel_id}"):
            st.session_state[f"edit_cli_{sel_id}"] = not st.session_state.get(f"edit_cli_{sel_id}", False)
            st.rerun()

        if st.button("🗑️ Cancella Cliente", use_container_width=True, key=f"ask_del_{sel_id}"):
            st.session_state["confirm_delete_cliente"] = sel_id
            st.rerun()

    # === CONFERMA CANCELLAZIONE ===
    if st.session_state.get("confirm_delete_cliente") == sel_id:
        st.warning(f"⚠️ Eliminare definitivamente **{cliente['RagioneSociale']}** (ID {sel_id}) e tutti i contratti associati?")
        cdel1, cdel2 = st.columns(2)
        with cdel1:
            if st.button("✅ Sì, elimina", use_container_width=True, key=f"do_del_{sel_id}"):
                try:
                    df_cli_new = df_cli[df_cli["ClienteID"].astype(str) != sel_id].copy()
                    df_ct_new  = df_ct[df_ct["ClienteID"].astype(str)  != sel_id].copy()

                    try:
                        conn = get_connection()
                        cur = conn.cursor()
                        cur.execute("DELETE FROM contratti_clienti WHERE ClienteID = %s", (sel_id,))
                        cur.execute("DELETE FROM clienti WHERE ClienteID = %s", (sel_id,))
                        conn.commit()
                        conn.close()
                        st.success("🗑️ Cliente e contratti eliminati da MySQL.")
                    except Exception as e:
                        st.error(f"⚠️ Errore eliminazione su MySQL: {e}")
                        df_cli_new.to_csv(CLIENTI_CSV, index=False, encoding="utf-8-sig")
                        df_ct_new.to_csv(CONTRATTI_CSV, index=False, encoding="utf-8-sig")
                        st.info("💾 Backup locale su CSV aggiornato.")

                    try:
                        st.cache_data.clear()
                    except:
                        pass
                    st.session_state.pop("confirm_delete_cliente", None)
                    time.sleep(0.5)
                    st.rerun()

                except Exception as e:
                    st.error(f"❌ Errore durante l'eliminazione: {e}")

        with cdel2:
            if st.button("❌ Annulla", use_container_width=True, key=f"undo_del_{sel_id}"):
                st.session_state.pop("confirm_delete_cliente", None)
                st.info("Operazione annullata.")
                st.rerun()

    # === ANAGRAFICA CLIENTE (visuale compatta tipo scheda) ===
    st.divider()
    st.markdown("### 🧾 Anagrafica Cliente")

    st.markdown(
        f"""
        <div style='font-size:15px; line-height:1.8; padding:10px 15px; background-color:#f8fafc;
                    border:1px solid #e5e7eb; border-radius:10px;'>
            <b>📍 Indirizzo:</b> {cliente.get('Indirizzo','')}<br>
            <b>🏙️ Città:</b> {cliente.get('Citta','')} &nbsp;&nbsp; <b>📮 CAP:</b> {cliente.get('CAP','')}<br>
            <b>📞 Telefono:</b> {cliente.get('Telefono','')} &nbsp;&nbsp; <b>📱 Cellulare:</b> {cliente.get('Cell','')}<br>
            <b>✉️ Email:</b> {cliente.get('Email','')}<br>
            <b>👤 Referente:</b> {cliente.get('PersonaRiferimento','')}<br>
            <b>💼 Partita IVA:</b> {cliente.get('PartitaIVA','')}<br>
            <b>🏦 IBAN:</b> {cliente.get('IBAN','')} &nbsp;&nbsp; <b>📡 SDI:</b> {cliente.get('SDI','')}<br>
            <b>🧭 TMK:</b> {cliente.get('TMK','')}
        </div>
        """,
        unsafe_allow_html=True
    )

    # === EDIT ANAGRAFICA ===
    if st.session_state.get(f"edit_cli_{sel_id}", False):
        st.info("✏️ Modifica anagrafica attiva")
        with st.form(f"frm_anagrafica_{sel_id}"):
            c1, c2 = st.columns(2)
            with c1:
                indirizzo = st.text_input("📍 Indirizzo", cliente.get("Indirizzo", ""))
                citta     = st.text_input("🏙️ Città", cliente.get("Citta", ""))
                cap       = st.text_input("📮 CAP", cliente.get("CAP", ""))
                telefono  = st.text_input("📞 Telefono", cliente.get("Telefono", ""))
                cell      = st.text_input("📱 Cellulare", cliente.get("Cell", ""))
            with c2:
                email     = st.text_input("✉️ Email", cliente.get("Email", ""))
                persona   = st.text_input("👤 Persona Riferimento", cliente.get("PersonaRiferimento", ""))
                piva      = st.text_input("💼 Partita IVA", cliente.get("PartitaIVA", ""))
                iban      = st.text_input("🏦 IBAN", cliente.get("IBAN", ""))
                sdi       = st.text_input("📡 SDI", cliente.get("SDI", ""))

            tmk_options = sorted(["Giulia", "Antonella", "Laura", "Annalisa"])
            tmk_attuale = cliente.get("TMK", "")
            tmk_sel = st.selectbox("🧭 Assegna TMK", tmk_options, index=tmk_options.index(tmk_attuale) if tmk_attuale in tmk_options else 0)

            salva = st.form_submit_button("💾 Salva Modifiche")
            if salva:
                try:
                    idx = df_cli.index[df_cli["ClienteID"].astype(str) == sel_id][0]
                    df_cli.loc[idx, [
                        "Indirizzo","Citta","CAP","Telefono","Cell","Email",
                        "PersonaRiferimento","PartitaIVA","IBAN","SDI","TMK"
                    ]] = [indirizzo, citta, cap, telefono, cell, email, persona, piva, iban, sdi, tmk_sel]

                    try:
                        save_table(df_cli, "clienti")
                        st.success("✅ Anagrafica aggiornata su MySQL!")
                    except Exception as e:
                        st.error(f"⚠️ Errore salvataggio MySQL: {e}")
                        save_clienti(df_cli)
                        st.info("💾 Backup locale su CSV eseguito.")

                    st.session_state[f"edit_cli_{sel_id}"] = False
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Errore durante il salvataggio: {e}")

    # === NOTE CLIENTE ===
    st.divider()
    st.markdown("### 📝 Note Cliente")
    st.caption("Annotazioni o informazioni utili sul cliente (visibili a tutti gli utenti).")

    note_attuali = cliente.get("NoteCliente", "")
    nuove_note = st.text_area("Scrivi o modifica le note del cliente:", note_attuali, height=160, key=f"note_{sel_id}_{int(time.time()*1000)}")

    n1, n2 = st.columns([0.25, 0.75])
    with n1:
        if st.button("💾 Salva Note", use_container_width=True, key=f"save_note_{sel_id}"):
            try:
                idx_row = df_cli.index[df_cli["ClienteID"].astype(str) == sel_id][0]
                df_cli.loc[idx_row, "NoteCliente"] = nuove_note

                try:
                    save_table(df_cli, "clienti")
                    st.success("✅ Note salvate su MySQL!")
                except Exception as e:
                    st.error(f"⚠️ Errore salvataggio MySQL: {e}")
                    save_clienti(df_cli)
                    st.info("💾 Backup locale su CSV eseguito.")

                st.rerun()
            except Exception as e:
                st.error(f"❌ Errore durante il salvataggio: {e}")
    with n2:
        st.info("Le modifiche vengono salvate anche in MySQL (con backup CSV).")

    # === RECALL & VISITE ===
    st.divider()
    st.markdown("### ⚡ Recall e Visite")

    def _safe_date(val):
        try:
            d = pd.to_datetime(val, dayfirst=True)
            return None if pd.isna(d) else d.date()
        except Exception:
            return None

    ur_val = _safe_date(cliente.get("UltimoRecall"))
    pr_val = _safe_date(cliente.get("ProssimoRecall"))
    uv_val = _safe_date(cliente.get("UltimaVisita"))
    pv_val = _safe_date(cliente.get("ProssimaVisita"))

    if ur_val and not pr_val:
        pr_val = (pd.Timestamp(ur_val) + pd.DateOffset(months=3)).date()
    if uv_val and not pv_val:
        pv_val = (pd.Timestamp(uv_val) + pd.DateOffset(months=6)).date()

    uniq = f"{sel_id}_{int(time.time()*1000)}"
    r1, r2, r3, r4 = st.columns(4)
    ur = r1.date_input("⏰ Ultimo Recall",  value=ur_val, format="DD/MM/YYYY", key=f"ur_{uniq}")
    pr = r2.date_input("📅 Prossimo Recall", value=pr_val, format="DD/MM/YYYY", key=f"pr_{uniq}")
    uv = r3.date_input("👣 Ultima Visita",  value=uv_val, format="DD/MM/YYYY", key=f"uv_{uniq}")
    pv = r4.date_input("🗓️ Prossima Visita", value=pv_val, format="DD/MM/YYYY", key=f"pv_{uniq}")

    if st.button("💾 Salva Aggiornamenti", use_container_width=True, key=f"save_recall_{uniq}"):
        try:
            idx = df_cli.index[df_cli["ClienteID"].astype(str) == sel_id][0]
            df_cli.loc[idx, ["UltimoRecall","ProssimoRecall","UltimaVisita","ProssimaVisita"]] = \
                [fmt_date(ur), fmt_date(pr), fmt_date(uv), fmt_date(pv)]

            try:
                save_table(df_cli, "clienti")
                st.success("✅ Date aggiornate su MySQL!")
            except Exception as e:
                st.error(f"⚠️ Errore salvataggio MySQL: {e}")
                save_clienti(df_cli)
                st.info("💾 Backup locale su CSV eseguito.")

            st.rerun()
        except Exception as e:
            st.error(f"❌ Errore salvataggio recall/visite: {e}")

    # === GESTIONE PREVENTIVI ===
    st.divider()
    st.markdown("### 🧾 Gestione Preventivi")

    TEMPLATE_OPTIONS = {
        "Offerta A4": "Offerta_A4.docx",
        "Offerta A3": "Offerta_A3.docx",
        "Centralino": "Offerta_Centralino.docx",
        "Varie": "Offerta_Varie.docx",
    }
    PREVENTIVI_DIR = STORAGE_DIR / "preventivi"
    PREVENTIVI_DIR.mkdir(parents=True, exist_ok=True)
    prev_csv = STORAGE_DIR / "preventivi.csv"

    if prev_csv.exists():
        df_prev = pd.read_csv(prev_csv, dtype=str).fillna("")
    else:
        df_prev = pd.DataFrame(columns=["ClienteID", "NumeroOfferta", "Template", "NomeFile", "Percorso", "DataCreazione"])

    # === CREA NUOVO PREVENTIVO ===
    st.markdown("#### ➕ Crea nuovo preventivo")

    anno = datetime.now().year
    nome_cliente = cliente.get("RagioneSociale", "")
    nome_sicuro = "".join(c for c in nome_cliente if c.isalnum())[:6].upper()
    num_off = f"OFF-{anno}-{nome_sicuro}-{len(df_prev[df_prev['ClienteID'].astype(str) == sel_id]) + 1:03d}"

    with st.form(f"frm_prev_{sel_id}"):
        st.text_input("Numero Offerta", num_off, disabled=True)
        nome_file = st.text_input("Nome File", f"{num_off}.docx")
        template = st.selectbox("Template", list(TEMPLATE_OPTIONS.keys()))
        genera_btn = st.form_submit_button("💾 Genera Preventivo")

    if genera_btn:
        try:
            from docx import Document
            tpl_path = STORAGE_DIR / "templates" / TEMPLATE_OPTIONS[template]

            if not tpl_path.exists():
                st.error(f"❌ Template non trovato: {tpl_path}")
                st.stop()

            doc = Document(tpl_path)
            mappa = {
                "CLIENTE": nome_cliente,
                "INDIRIZZO": cliente.get("Indirizzo", ""),
                "CITTA": cliente.get("Citta", ""),
                "NUMERO_OFFERTA": num_off,
                "DATA": datetime.now().strftime("%d/%m/%Y"),
                "ULTIMO_RECALL": fmt_date(cliente.get("UltimoRecall")),
                "PROSSIMO_RECALL": fmt_date(cliente.get("ProssimoRecall")),
                "ULTIMA_VISITA": fmt_date(cliente.get("UltimaVisita")),
                "PROSSIMA_VISITA": fmt_date(cliente.get("ProssimaVisita")),
            }

            for p in doc.paragraphs:
                for k, v in mappa.items():
                    if f"<<{k}>>" in p.text:
                        for run in p.runs:
                            run.text = run.text.replace(f"<<{k}>>", str(v))

            out_path = PREVENTIVI_DIR / nome_file
            doc.save(out_path)
            box_upload_if_changed(out_path)

            nuova_riga = {
                "ClienteID": sel_id,
                "NumeroOfferta": num_off,
                "Template": TEMPLATE_OPTIONS[template],
                "NomeFile": nome_file,
                "Percorso": str(out_path),
                "DataCreazione": datetime.now().strftime("%d/%m/%Y %H:%M"),
            }

            # ✅ Evita duplicati
            df_prev = pd.concat([df_prev, pd.DataFrame([nuova_riga])], ignore_index=True)
            df_prev = df_prev.drop_duplicates(subset=["ClienteID", "NomeFile"], keep="last")

            df_prev.to_csv(prev_csv, index=False, encoding="utf-8-sig")

            st.success(f"✅ Preventivo generato: {out_path.name}")
            st.rerun()
        except Exception as e:
            import traceback
            st.error(f"❌ Errore durante la generazione del preventivo:\n\n{traceback.format_exc()}")

    # === ELENCO PREVENTIVI DEL CLIENTE ===
    st.divider()
    st.markdown("#### 📂 Elenco Preventivi Cliente")

    df_prev_cli = df_prev[df_prev["ClienteID"].astype(str) == str(sel_id)]
    if not df_prev_cli.empty:
        for i, row in df_prev_cli.iterrows():
            nome_file = row["NomeFile"]
            percorso = row["Percorso"]
            data = row.get("DataCreazione", "")
            col1, col2, col3 = st.columns([4, 1, 1])
            col1.write(f"📄 **{nome_file}**  \n🕓 {data}")

            if Path(percorso).exists():
                with open(percorso, "rb") as f:
                    col2.download_button("⬇️", f, file_name=nome_file,
                                         mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document")
            else:
                col2.warning("❌ File mancante")

            if col3.button("🗑️", key=f"del_prev_{i}", help="Elimina preventivo"):
                try:
                    df_prev = df_prev.drop(index=i)
                    df_prev.to_csv(prev_csv, index=False, encoding="utf-8-sig")
                    if Path(percorso).exists():
                        Path(percorso).unlink()
                    st.success(f"✅ Preventivo '{nome_file}' eliminato.")
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Errore eliminazione: {e}")
    else:
        st.info("Nessun preventivo per questo cliente.")

# =====================================
# PAGINA CONTRATTI — VERSIONE STABILE 2025 (senza duplicati widget)
# =====================================
def page_contratti(df_cli: pd.DataFrame, df_ct: pd.DataFrame, role: str):
    # FIX: sincronizza selezione cliente se arriviamo da pulsante esterno
    if "selected_cliente" in st.session_state:
        selected_id = st.session_state.pop("selected_cliente")
        df_cli["label"] = df_cli.apply(lambda r: f"{r['ClienteID']} — {r['RagioneSociale']}", axis=1)
        label = df_cli[df_cli["ClienteID"] == selected_id]["label"].values
        if len(label) > 0:
            st.session_state["sel_cli_ct"] = label[0]

    ruolo_scrittura = st.session_state.get("ruolo_scrittura", role)
    permessi_limitati = ruolo_scrittura == "limitato"

    st.markdown("<h2>📄 Gestione Contratti</h2>", unsafe_allow_html=True)
    st.divider()

    # === Selezione Cliente ===
    if df_cli.empty:
        st.info("Nessun cliente presente.")
        return

    labels = df_cli["RagioneSociale"].astype(str).tolist()
    # === Selectbox clienti con controllo di coerenza ===
    if "sel_cli_ct" in st.session_state and st.session_state["sel_cli_ct"] not in labels:
        st.session_state.pop("sel_cli_ct")  # reset automatico se valore non più valido
    
    sel_label = st.selectbox(
        "Seleziona Cliente",
        labels,
        index=0 if labels else None,
        key="sel_cli_ct"
    )

    
    # Quando l'utente seleziona un cliente
    if sel_label:
        sel_id = df_cli.loc[df_cli["RagioneSociale"] == sel_label, "ClienteID"].iloc[0]
    
        # 🔹 Pulsante per aprire e poi azzerare la selezione
        if st.button("📂 Apri Cliente Selezionato"):
            st.session_state["selected_cliente"] = sel_id
            # 🔁 Reset del campo selectbox
            st.session_state.pop("sel_cli_ct", None)
            st.rerun()

    rag_soc = df_cli.loc[df_cli["ClienteID"] == sel_id, "RagioneSociale"].iloc[0]

    st.markdown(f"<h3 style='text-align:center;color:#2563eb'>{rag_soc}</h3>", unsafe_allow_html=True)
    st.caption(f"ID Cliente: {sel_id}")

    # === Filtra contratti del cliente ===
    ct = df_ct[df_ct["ClienteID"].astype(str) == str(sel_id)].copy()
    if not ct.empty:
        ct = ct[
            (ct["NumeroContratto"].astype(str).str.strip() != "") |
            (ct["DescrizioneProdotto"].astype(str).str.strip() != "")
        ]
        ct = ct.dropna(how="all").reset_index(drop=True)

    # === CREA NUOVO CONTRATTO ===
    with st.expander("➕ Crea Nuovo Contratto", expanded=False):
        if permessi_limitati:
            st.warning("🔒 Accesso sola lettura")
        else:
            with st.form(f"new_ct_{sel_id}"):
                st.markdown("#### 📄 Dati Contratto")
                c1, c2, c3, c4 = st.columns(4)
                num = c1.text_input("Numero Contratto")
                din = c2.date_input("Data Inizio", format="DD/MM/YYYY")
                durata = c3.selectbox("Durata (mesi)", DURATE_MESI, index=2)
                stato_new = c4.selectbox("Stato", ["aperto", "chiuso"], index=0)
                desc = st.text_area("Descrizione Prodotto", height=80)

                st.markdown("#### 💰 Dati Economici")
                c5, c6, c7 = st.columns(3)
                nf = c5.text_input("NOL_FIN")
                ni = c6.text_input("NOL_INT")
                tot = c7.text_input("TotRata")

                st.markdown("#### 🖨️ Copie incluse ed Eccedenze")
                c8, c9, c10, c11 = st.columns(4)
                copie_bn = c8.text_input("Copie incluse B/N", value="")
                ecc_bn = c9.text_input("Eccedenza B/N (€)", value="")
                copie_col = c10.text_input("Copie incluse Colore", value="")
                ecc_col = c11.text_input("Eccedenza Colore (€)", value="")

                if st.form_submit_button("💾 Crea contratto"):
                    try:
                        fine = pd.to_datetime(din) + pd.DateOffset(months=int(durata))
                        nuovo = {
                            "ClienteID": sel_id,
                            "RagioneSociale": rag_soc,
                            "NumeroContratto": num,
                            "DataInizio": fmt_date(din),
                            "DataFine": fmt_date(fine),
                            "Durata": durata,
                            "DescrizioneProdotto": desc,
                            "NOL_FIN": nf,
                            "NOL_INT": ni,
                            "TotRata": tot,
                            "CopieBN": copie_bn,
                            "EccBN": ecc_bn,
                            "CopieCol": copie_col,
                            "EccCol": ecc_col,
                            "Stato": stato_new
                        }
                        if not num.strip() and not desc.strip():
                            st.warning("⚠️ Inserisci almeno il numero contratto o una descrizione valida.")
                        else:
                            df_ct = pd.concat([df_ct, pd.DataFrame([nuovo])], ignore_index=True)
                            df_ct = df_ct.dropna(how="all").reset_index(drop=True)
                            save_contratti(df_ct)
                            st.success("✅ Contratto creato correttamente.")
                            st.rerun()
                    except Exception as e:
                        st.error(f"❌ Errore durante la creazione del contratto: {e}")

    # === STILE TABELLA ===
    st.markdown("""
    <style>
      .tbl-wrap{border:1px solid #e5e7eb; border-radius:12px; overflow:hidden;
        box-shadow:0 4px 16px rgba(0,0,0,.06); background:#fff;}
      .tbl-head{background:#2563eb; color:#fff; font-weight:700; font-size:13px;}
      .tbl-head > div{padding:10px 8px; text-align:center; border-right:1px solid rgba(255,255,255,.25);}
      .tbl-row{font-size:13px; border-top:1px solid #eef2f7;}
      .tbl-row:hover{background:#f6faff;}
      .cell{padding:8px 8px; text-align:center;}
      .cell-left{text-align:left;}
      .row-closed{background:#fff1f1 !important;}
    </style>
    """, unsafe_allow_html=True)

    st.markdown("### 📋 Contratti del Cliente")

    if ct.empty:
        st.info("Nessun contratto registrato.")
        return

    # --- intestazione
    head_cols = st.columns([0.7, 0.9, 0.9, 0.8, 2.8, 1.1, 1, 1, 0.9, 0.9, 0.9, 0.9, 1])
    head_lbls = ["N°","Inizio","Fine","Durata","Descrizione Prodotto","Tot. Rata","NOL FIN","NOL INT","Copie B/N","Ecc. B/N","Copie Col","Ecc. Col","Azioni"]
    for c, h in zip(head_cols, head_lbls):
        c.markdown(f"<div class='tbl-head'><div>{h}</div></div>", unsafe_allow_html=True)

    # --- righe
    oggi = pd.Timestamp.now().normalize()
    for i, r in ct.iterrows():
        stato = str(r.get("Stato","aperto")).lower()
        row_cls = "row-closed" if stato == "chiuso" else ""

        cols = st.columns([0.7, 0.9, 0.9, 0.8, 2.8, 1.1, 1, 1, 0.9, 0.9, 0.9, 0.9, 1])
        cols[0].markdown(f"<div class='tbl-row {row_cls}'><div class='cell mono'>{r.get('NumeroContratto','—')}</div></div>", unsafe_allow_html=True)
        cols[1].markdown(f"<div class='tbl-row {row_cls}'><div class='cell mono'>{fmt_date(r.get('DataInizio'))}</div></div>", unsafe_allow_html=True)
        cols[2].markdown(f"<div class='tbl-row {row_cls}'><div class='cell mono'>{fmt_date(r.get('DataFine'))}</div></div>", unsafe_allow_html=True)
        cols[3].markdown(f"<div class='tbl-row {row_cls}'><div class='cell'>{safe_text(r.get('Durata',''))}</div></div>", unsafe_allow_html=True)
        cols[4].markdown(f"<div class='tbl-row {row_cls}'><div class='cell cell-left'>{safe_text(r.get('DescrizioneProdotto',''))}</div></div>", unsafe_allow_html=True)
        cols[5].markdown(f"<div class='tbl-row {row_cls}'><div class='cell cell-right'>{money(r.get('TotRata')) or '—'}</div></div>", unsafe_allow_html=True)
        cols[6].markdown(f"<div class='tbl-row {row_cls}'><div class='cell mono'>{r.get('NOL_FIN','')}</div></div>", unsafe_allow_html=True)
        cols[7].markdown(f"<div class='tbl-row {row_cls}'><div class='cell mono'>{r.get('NOL_INT','')}</div></div>", unsafe_allow_html=True)
        cols[8].markdown(f"<div class='tbl-row {row_cls}'><div class='cell mono'>{r.get('CopieBN','')}</div></div>", unsafe_allow_html=True)
        cols[9].markdown(f"<div class='tbl-row {row_cls}'><div class='cell mono'>{r.get('EccBN','')}</div></div>", unsafe_allow_html=True)
        cols[10].markdown(f"<div class='tbl-row {row_cls}'><div class='cell mono'>{r.get('CopieCol','')}</div></div>", unsafe_allow_html=True)
        cols[11].markdown(f"<div class='tbl-row {row_cls}'><div class='cell mono'>{r.get('EccCol','')}</div></div>", unsafe_allow_html=True)

        # --- azioni (chiavi univoche)
        # --- azioni (chiavi univoche)
        with cols[12]:
            b1, b2, b3 = st.columns(3)

            # ✏️ Modifica contratto → apre la pagina dedicata
            if b1.button("✏️", key=f"edit_ct_{i}", help="Modifica contratto", disabled=permessi_limitati):
                st.session_state["edit_gidx"] = i
                st.session_state["nav_target"] = "✏️ Modifica Contratto"
                st.rerun()

            # 🔒 Chiudi / Riapri contratto
            stato_btn = "🔒" if stato != "chiuso" else "🟢"
            if b2.button(stato_btn, key=f"lock_ct_{i}", help="Chiudi/Riapri contratto", disabled=permessi_limitati):
                try:
                    nuovo_stato = "chiuso" if stato != "chiuso" else "aperto"
                    df_ct.loc[df_ct.index == i, "Stato"] = nuovo_stato

                    try:
                        save_table(df_ct, "contratti_clienti")
                        st.toast(f"🔁 Stato contratto aggiornato su MySQL: {nuovo_stato.upper()}", icon="✅")
                    except Exception as e:
                        st.error(f"⚠️ Errore salvataggio MySQL: {e}")
                        save_contratti(df_ct)
                        st.info("💾 Backup locale su CSV eseguito.")

                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Errore durante l’aggiornamento dello stato: {e}")

            # 🗑️ Elimina contratto
            if b3.button("🗑️", key=f"del_ct_{i}", help="Elimina contratto", disabled=permessi_limitati):
                st.session_state["delete_gidx"] = i
                st.session_state["ask_delete_now"] = True
                st.rerun()

    # === ELIMINAZIONE CONTRATTO (MySQL + CSV fallback) ===
    if st.session_state.get("ask_delete_now") and st.session_state.get("delete_gidx") is not None:
        gidx = st.session_state["delete_gidx"]
        if gidx in ct.index:
            contratto = ct.loc[gidx]
            numero = contratto.get("NumeroContratto", "Senza numero")
            contratto_id = str(contratto.get("ContrattoID", ""))

            st.warning(f"⚠️ Eliminare definitivamente il contratto **{numero}**?")
            c1, c2 = st.columns(2)
            with c1:
                if st.button("✅ Sì, elimina", use_container_width=True):
                    try:
                        # 🔹 Rimuovi contratto dal DataFrame
                        df_ct_new = df_ct.drop(index=gidx).reset_index(drop=True)

                        # 🔹 Prova a eliminare anche da MySQL
                        try:
                            conn = get_connection()
                            cur = conn.cursor()
                            if contratto_id:
                                cur.execute("DELETE FROM contratti_clienti WHERE ContrattoID = %s", (contratto_id,))
                            else:
                                cur.execute("DELETE FROM contratti_clienti WHERE NumeroContratto = %s", (numero,))
                            conn.commit()
                            conn.close()
                            st.success("🗑️ Contratto eliminato da MySQL.")
                        except Exception as e:
                            st.error(f"⚠️ Errore eliminazione su MySQL: {e}")
                            save_contratti(df_ct_new)
                            st.info("💾 Backup locale su CSV aggiornato.")

                        # 🔹 Aggiorna cache e interfaccia
                        try:
                            st.cache_data.clear()
                        except:
                            pass

                        df_ct = df_ct_new
                        st.session_state.pop("ask_delete_now", None)
                        st.session_state.pop("delete_gidx", None)
                        time.sleep(0.5)
                        st.rerun()

                    except Exception as e:
                        st.error(f"❌ Errore durante l'eliminazione: {e}")

            with c2:
                if st.button("❌ Annulla", use_container_width=True):
                    st.session_state.pop("ask_delete_now", None)
                    st.session_state.pop("delete_gidx", None)
                    st.info("Operazione annullata.")
                    st.rerun()



    # === ESPORTAZIONI (Excel + PDF) ===
    st.divider()
    st.markdown("### 📤 Esportazioni")

    from datetime import datetime
    data_export = datetime.now().strftime("%d/%m/%Y")

    # === EXPORT EXCEL ===
    cex1, cex2 = st.columns(2)
    with cex1:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from io import BytesIO

            wb = Workbook()
            ws = wb.active
            ws.title = f"Contratti {rag_soc}"

            # 🔹 Titolo
            ws.merge_cells("A1:L1")
            cell_title = ws["A1"]
            cell_title.value = f"Contratti Cliente: {rag_soc} — Data: {data_export}"
            cell_title.font = Font(bold=True, size=14)
            cell_title.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 25

            # 🔹 Intestazioni coerenti
            headers = [
                "N°", "Inizio", "Fine", "Durata", "Descrizione Prodotto",
                "Tot. Rata", "NOL FIN", "NOL INT", "Copie B/N",
                "Ecc. B/N", "Copie Col", "Ecc. Col"
            ]
            ws.append(headers)

            yellow_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
            header_font = Font(bold=True, color="000000")
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

            for col_idx, header in enumerate(headers, 1):
                c = ws.cell(row=2, column=col_idx)
                c.fill = yellow_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = thin_border
                ws.column_dimensions[get_column_letter(col_idx)].width = 18

            # 🔹 Righe dati
            for _, r in ct.iterrows():
                ws.append([
                    r.get("NumeroContratto", ""),
                    fmt_date(r.get("DataInizio")),
                    fmt_date(r.get("DataFine")),
                    r.get("Durata", ""),
                    r.get("DescrizioneProdotto", ""),
                    money(r.get("TotRata")),
                    r.get("NOL_FIN", ""),
                    r.get("NOL_INT", ""),
                    r.get("CopieBN", ""),
                    r.get("EccBN", ""),
                    r.get("CopieCol", ""),
                    r.get("EccCol", "")
                ])

            # 🔹 Allineamento e altezza righe dinamica
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=12):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                    cell.border = thin_border
                ws.row_dimensions[row[0].row].height = 22 + (len(str(row[4].value)) // 70) * 10

            bio = BytesIO()
            wb.save(bio)
            st.download_button(
                "📘 Esporta Excel",
                bio.getvalue(),
                file_name=f"Contratti_{rag_soc}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Errore export Excel: {e}")

        # === EXPORT PDF (centrato in pagina, A4 orizzontale, 1 pagina quando possibile) ===
    with cex2:
        try:
            from fpdf import FPDF
            import requests
            from io import BytesIO

            LOGO_URL = "https://www.shtsrl.com/template/images/logo.png"

            pdf = FPDF("L", "mm", "A4")
            # Disabilito l'autobreak per evitare pagine “vuote” solo col footer
            pdf.set_auto_page_break(auto=False)
            pdf.add_page()

            # Margini più “grafici”
            left_margin = 12
            right_margin = 12
            top_margin = 10
            bottom_margin = 12
            pdf.set_margins(left=left_margin, top=top_margin, right=right_margin)

            page_w = pdf.w
            usable_w = page_w - left_margin - right_margin

            # === Logo SHT centrato ===
            try:
                resp = requests.get(LOGO_URL, timeout=5)
                if resp.status_code == 200:
                    logo_bytes = BytesIO(resp.content)
                    # logo di 35mm, centrato
                    logo_w = 35
                    x_logo = left_margin + (usable_w - logo_w) / 2.0
                    pdf.image(logo_bytes, x=x_logo, y=8, w=logo_w)
            except Exception:
                pass

            # Spazio sotto il logo
            pdf.set_y(8 + 35 + 4)

            # === Titolo centrato ===
            pdf.set_font("Arial", "B", 13)
            pdf.cell(0, 8, safe_text(f"Contratti Cliente: {rag_soc} - {data_export}"), ln=1, align="C")
            pdf.ln(3)

            # === Intestazioni + larghezze colonna ===
            headers = [
                "N°", "Inizio", "Fine", "Durata", "Descrizione Prodotto",
                "Tot. Rata", "NOL FIN", "NOL INT", "Copie B/N",
                "Ecc. B/N", "Copie Col", "Ecc. Col"
            ]
            # Larghezze pensate per A4 orizzontale, ma calcoliamo lo start centrato
            col_widths = [10, 20, 20, 15, 110, 25, 20, 20, 22, 22, 22, 22]
            table_w = sum(col_widths)
            # Se la tabella è più larga dello spazio utile, la riduciamo in scala uniforme
            if table_w > usable_w:
                scale = usable_w / table_w
                col_widths = [w * scale for w in col_widths]
                table_w = usable_w

            # X iniziale per centrare
            start_x = left_margin + (usable_w - table_w) / 2.0

            # === Header riga ===
            pdf.set_font("Arial", "B", 9)
            pdf.set_fill_color(255, 253, 231)
            pdf.set_xy(start_x, pdf.get_y())
            for h, w in zip(headers, col_widths):
                pdf.cell(w, 8, safe_text(h), border=1, align="C", fill=True)
            pdf.ln(8)

            # === Dati ===
            pdf.set_font("Arial", "", 8)
            row_gap = 0  # nessun gap extra, per comprimere in una pagina quando possibile

            # Altezza minima riga e gestione wrap descrizione
            def draw_row(r):
                nonlocal start_x
                stato = str(r.get("Stato", "aperto")).lower()
                is_closed = (stato == "chiuso")
                fill_color = (255, 230, 230) if is_closed else (255, 255, 255)

                row_values = [
                    r.get("NumeroContratto", ""),
                    fmt_date(r.get("DataInizio")),
                    fmt_date(r.get("DataFine")),
                    r.get("Durata", ""),
                    safe_text(r.get("DescrizioneProdotto", "")),
                    money(r.get("TotRata")),
                    r.get("NOL_FIN", ""),
                    r.get("NOL_INT", ""),
                    r.get("CopieBN", ""),
                    r.get("EccBN", ""),
                    r.get("CopieCol", ""),
                    r.get("EccCol", "")
                ]

                # Calcola quante “linee” servono per la descrizione stimando ~95 char per 110mm (scala se ridotta)
                # Più robusto: usa la larghezza reale del font
                pdf.set_font("Arial", "", 8)
                desc_text = str(row_values[4])
                desc_w = col_widths[4]
                # stimiamo quante righe servono alla descrizione con la funzione di misura stringa
                # dividendo in base alla larghezza disponibile
                words = desc_text.split()
                lines = []
                line = ""
                for w in words:
                    test = (line + " " + w).strip()
                    if pdf.get_string_width(test) <= (desc_w - 2):  # un pochino di padding
                        line = test
                    else:
                        lines.append(line)
                        line = w
                if line:
                    lines.append(line)
                desc_lines = max(1, len(lines))

                line_h = 5.5  # un po’ compatto per favorire “una pagina”
                row_h = max(6, line_h * desc_lines)

                # Se sforza il fondo (considero 10 mm di margine + 6 di footer)
                bottom_limit = pdf.h - bottom_margin - 6
                if pdf.get_y() + row_h > bottom_limit:
                    pdf.add_page()
                    pdf.set_xy(start_x, top_margin + 12)  # spazio per allineare con titolo mancante
                    # re-disegno l’header su nuova pagina
                    pdf.set_font("Arial", "B", 9)
                    pdf.set_fill_color(255, 253, 231)
                    for h, w in zip(headers, col_widths):
                        pdf.cell(w, 8, safe_text(h), border=1, align="C", fill=True)
                    pdf.ln(8)
                    pdf.set_font("Arial", "", 8)

                # Disegna la riga
                pdf.set_fill_color(*fill_color)
                x0 = start_x
                y0 = pdf.get_y()
                for idx, (val, wcol) in enumerate(zip(row_values, col_widths)):
                    pdf.set_xy(x0, y0)
                    if idx == 4:
                        # descrizione: multicell con bordo, poi mi riposiziono
                        pdf.multi_cell(wcol, line_h, safe_text(str(val)), border=1, align="L", fill=(fill_color != (255, 255, 255)))
                        # calcola dove ripartire: max(y raggiunta, y0+row_h)
                        y_after = pdf.get_y()
                        # posiziona x subito dopo la colonna e ripristina y per le prossime celle
                        pdf.set_xy(x0 + wcol, y0)
                    else:
                        pdf.cell(wcol, row_h, safe_text(str(val)), border=1, align="C", fill=(fill_color != (255, 255, 255)))
                    x0 += wcol
                pdf.ln(row_h + row_gap)

            for _, r in ct.iterrows():
                draw_row(r)

            # === Footer centrato ===
            pdf.set_text_color(100, 100, 100)
            pdf.set_font("Arial", "I", 8)
            pdf.set_y(pdf.h - bottom_margin)
            pdf.cell(0, 6, safe_text("SHT S.r.l. - Tutti i diritti riservati"), 0, 0, "C")

            pdf_bytes = pdf.output(dest="S").encode("latin-1", errors="replace")
            st.download_button(
                "📗 Esporta PDF",
                pdf_bytes,
                file_name=f"Contratti_{rag_soc}.pdf",
                mime="application/pdf"
            )

        except Exception as e:
            st.error(f"Errore export PDF: {e}")

# =====================================
# PAGINA DI MODIFICA CONTRATTO (VERSIONE CORRETTA)
# =====================================
def page_modifica_contratto(df_cli: pd.DataFrame, df_ct: pd.DataFrame, role: str):
    """Pagina dedicata alla modifica di un contratto selezionato"""
    if "edit_gidx" not in st.session_state:
        st.warning("⚠️ Nessun contratto selezionato per la modifica.")
        st.stop()

    gidx = st.session_state["edit_gidx"]
    if gidx not in df_ct.index:
        st.error("❌ Contratto non trovato.")
        st.stop()

    contratto = df_ct.loc[gidx]
    rag_soc = contratto.get("RagioneSociale", "—")
    st.markdown(f"## ✏️ Modifica Contratto — {rag_soc}")
    st.caption(f"Numero Contratto: {contratto.get('NumeroContratto','')}")

    st.divider()

    # Conversione sicura delle date
    din_val = pd.to_datetime(contratto.get("DataInizio"), dayfirst=True, errors="coerce")
    dfi_val = pd.to_datetime(contratto.get("DataFine"), dayfirst=True, errors="coerce")

    din_default = din_val if pd.notna(din_val) else datetime.today()
    dfi_default = dfi_val if pd.notna(dfi_val) else datetime.today()

    # === FORM ===
    with st.form("frm_edit_contract_page"):
        col1, col2, col3 = st.columns(3)
        num = col1.text_input("📄 Numero Contratto", contratto.get("NumeroContratto", ""))
        durata = col2.text_input("📆 Durata (mesi)", contratto.get("Durata", ""))
        stato = col3.selectbox("📋 Stato", ["aperto", "chiuso"],
                               index=0 if contratto.get("Stato", "") != "chiuso" else 1)

        col4, col5 = st.columns(2)
        din = col4.date_input("📅 Data Inizio", value=din_default, format="DD/MM/YYYY")
        dfi = col5.date_input("📅 Data Fine", value=dfi_default, format="DD/MM/YYYY")

        desc = st.text_area("🧾 Descrizione Prodotto", contratto.get("DescrizioneProdotto", ""), height=100)

        st.markdown("### 💰 Dati Economici")
        e1, e2, e3 = st.columns(3)
        nf = e1.text_input("NOL_FIN", contratto.get("NOL_FIN", ""))
        ni = e2.text_input("NOL_INT", contratto.get("NOL_INT", ""))
        tot = e3.text_input("Tot. Rata (€)", contratto.get("TotRata", ""))

        st.markdown("### 🖨️ Copie incluse ed Eccedenze")
        c1, c2, c3, c4 = st.columns(4)
        copie_bn = c1.text_input("Copie B/N", contratto.get("CopieBN", ""))
        ecc_bn = c2.text_input("Eccedenza B/N (€)", contratto.get("EccBN", ""))
        copie_col = c3.text_input("Copie Colore", contratto.get("CopieCol", ""))
        ecc_col = c4.text_input("Eccedenza Colore (€)", contratto.get("EccCol", ""))

        # 🔹 Pulsanti (devono stare dentro il form!)
        col_s, col_a = st.columns(2)
        salva = col_s.form_submit_button("💾 Salva Modifiche")
        annulla = col_a.form_submit_button("❌ Annulla")

        if salva:
            try:
                df_ct.loc[gidx, [
                    "NumeroContratto", "DataInizio", "DataFine", "Durata",
                    "DescrizioneProdotto", "NOL_FIN", "NOL_INT", "TotRata",
                    "CopieBN", "EccBN", "CopieCol", "EccCol", "Stato"
                ]] = [
                    num, fmt_date(din), fmt_date(dfi), durata, desc,
                    nf, ni, tot, copie_bn, ecc_bn, copie_col, ecc_col, stato
                ]

                try:
                    save_table(df_ct, "contratti_clienti")
                    st.success("✅ Contratto salvato su MySQL!")
                except Exception as e:
                    st.error(f"⚠️ Errore salvataggio MySQL: {e}")
                    save_contratti(df_ct)
                    st.info("💾 Backup locale su CSV eseguito.")

                st.rerun()
            except Exception as e:
                st.error(f"❌ Errore durante il salvataggio del contratto: {e}")

        if annulla:
            st.info("Operazione annullata.")
            st.session_state["nav_target"] = "Contratti"
            st.rerun()


# =====================================
# FUNZIONE MODALE MODIFICA CONTRATTO
# =====================================
def show_contract_modal(contratto, df_ct, df_cli, rag_soc):
    """Mostra la finestra di modifica al centro schermo"""
    st.markdown("""
    <style>
    .modal-bg {
        position: fixed; top:0; left:0; width:100%; height:100%;
        background: rgba(0,0,0,0.4); z-index: 9998;
        display: flex; align-items: center; justify-content: center;
    }
    .modal-box {
        background: white; border-radius: 12px; width: 620px;
        padding: 1.8rem 2rem; box-shadow: 0 4px 18px rgba(0,0,0,0.25);
    }
    </style>
    <div class="modal-bg"><div class="modal-box">
    """, unsafe_allow_html=True)

    st.markdown(f"### ✏️ Modifica Contratto {contratto.get('NumeroContratto','')}")
    with st.form("frm_edit_contract"):
        col1, col2 = st.columns(2)
        with col1:
            num = st.text_input("Numero Contratto", contratto.get("NumeroContratto",""), disabled=True)
            din = st.date_input("Data Inizio", value=pd.to_datetime(contratto.get("DataInizio"), dayfirst=True, errors="coerce"))
            durata = st.text_input("Durata (mesi)", contratto.get("Durata",""))
            stato = st.selectbox("Stato", ["aperto", "chiuso"], index=0 if contratto.get("Stato","")!="chiuso" else 1)
        with col2:
            nf = st.text_input("NOL_FIN", contratto.get("NOL_FIN",""))
            ni = st.text_input("NOL_INT", contratto.get("NOL_INT",""))
            tot = st.text_input("Tot Rata", contratto.get("TotRata",""))
        desc = st.text_area("Descrizione Prodotto", contratto.get("DescrizioneProdotto",""), height=100)
        colA, colB, colC, colD = st.columns(4)
        copie_bn = colA.text_input("Copie B/N", contratto.get("CopieBN",""))
        ecc_bn   = colB.text_input("Extra B/N (€)", contratto.get("EccBN",""))
        copie_col= colC.text_input("Copie Colore", contratto.get("CopieCol",""))
        ecc_col  = colD.text_input("Extra Colore (€)", contratto.get("EccCol",""))

        salva = st.form_submit_button("💾 Salva modifiche", use_container_width=True)
        annulla = st.form_submit_button("❌ Annulla", use_container_width=True)

        if salva:
            try:
                idx = df_ct.index[df_ct["NumeroContratto"] == num][0]
                df_ct.loc[idx, [
                    "DataInizio","Durata","DescrizioneProdotto","NOL_FIN","NOL_INT",
                    "TotRata","CopieBN","EccBN","CopieCol","EccCol","Stato"
                ]] = [
                    fmt_date(din), durata, desc, nf, ni, tot, copie_bn, ecc_bn, copie_col, ecc_col, stato
                ]
                save_contratti(df_ct)
                st.success("✅ Contratto aggiornato con successo.")
                time.sleep(0.6)
                st.experimental_set_query_params()
                st.rerun()
            except Exception as e:
                st.error(f"❌ Errore durante il salvataggio: {e}")

        if annulla:
            st.experimental_set_query_params()
            st.rerun()

    st.markdown("</div></div>", unsafe_allow_html=True)


# =====================================
# FUNZIONI DI ESPORTAZIONE (Excel + PDF)
# =====================================
def export_excel_contratti(df_ct, sel_id, rag_soc):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    disp = df_ct[df_ct["ClienteID"].astype(str) == str(sel_id)].copy()
    disp["DataInizio"] = disp["DataInizio"].apply(fmt_date)
    disp["DataFine"] = disp["DataFine"].apply(fmt_date)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Contratti {rag_soc}"
    ws.merge_cells("A1:M1")
    title = ws["A1"]
    title.value = f"Contratti Cliente: {rag_soc}"
    title.font = Font(size=14, bold=True, color="2563EB")
    title.alignment = Alignment(horizontal="center")

    headers = ["NumeroContratto", "DataInizio", "DataFine", "Durata", "DescrizioneProdotto",
               "NOL_FIN", "NOL_INT", "TotRata", "CopieBN", "EccBN", "CopieCol", "EccCol", "Stato"]
    ws.append(headers)

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2563EB")
    center = Alignment(horizontal="center", wrap_text=True)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i)
        c.font = head_font; c.fill = head_fill; c.alignment = center; c.border = thin

    for _, row in disp.iterrows():
        ws.append([str(row.get(h, "")) for h in headers])
        stato = str(row.get("Stato", "")).lower()
        r_idx = ws.max_row
        for j in range(1, len(headers)+1):
            cell = ws.cell(row=r_idx, column=j)
            cell.alignment = center
            cell.border = thin
            if stato == "chiuso":
                cell.fill = PatternFill("solid", fgColor="FFCDD2")

    for i in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(i)].width = 25

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def export_pdf_contratti(df_ct, sel_id, rag_soc):
    from fpdf import FPDF
    disp = df_ct[df_ct["ClienteID"].astype(str) == str(sel_id)].copy()
    disp["DataInizio"] = disp["DataInizio"].apply(fmt_date)
    disp["DataFine"] = disp["DataFine"].apply(fmt_date)
    headers = ["NumeroContratto", "DataInizio", "DataFine", "Durata", "TotRata", "Stato"]
    widths = [30, 25, 25, 15, 25, 20]

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, safe_text(f"Contratti Cliente: {rag_soc}"), ln=1, align="C")
    pdf.set_font("Arial", "B", 10)
    for i, h in enumerate(headers):
        pdf.cell(widths[i], 8, safe_text(h), 1, 0, "C", True)
    pdf.ln()
    pdf.set_font("Arial", "", 9)
    for _, r in disp.iterrows():
        for i, h in enumerate(headers):
            stato = str(r.get("Stato", "")).lower()
            if stato == "chiuso":
                pdf.set_fill_color(255, 235, 238)
                pdf.cell(widths[i], 7, safe_text(r.get(h, "")), 1, 0, "C", fill=True)
            else:
                pdf.cell(widths[i], 7, safe_text(r.get(h, "")), 1, 0, "C")
        pdf.ln()
    return pdf.output(dest="S").encode("latin-1", errors="replace")
# =====================================
# 📈 DASHBOARD GRAFICI — priva di dipendenze extra
# =====================================
def _to_dt(s):
    return pd.to_datetime(s, errors="coerce", dayfirst=True)

def _to_float_eur(x):
    """Converte '1.234,56' o '1234.56 €' → 1234.56 (float). Vuoto se non numerico."""
    if pd.isna(x): return None
    t = str(x).strip()
    if not t: return None
    # rimuovi simboli e spazi, normalizza separatori
    t = t.replace("€", "").replace("EUR", "").replace(" ", "")
    # se formato italiano 1.234,56 → togli i punti mille e cambia la virgola
    if "," in t and t.count(",") == 1 and "." in t:
        t = t.replace(".", "").replace(",", ".")
    elif "," in t and t.count(",") == 1 and "." not in t:
        t = t.replace(",", ".")
    try:
        return float(t)
    except Exception:
        try:
            return float(pd.to_numeric(t, errors="coerce"))
        except Exception:
            return None

def page_dashboard_grafici(df_cli: pd.DataFrame, df_ct: pd.DataFrame, role: str):
    st.image(LOGO_URL, width=120)
    st.markdown("<h2>📈 Dashboard Grafici</h2>", unsafe_allow_html=True)
    st.caption("Panoramica interattiva su clienti e contratti (filtri in alto).")
    st.divider()

    # ======== PREPARAZIONE DATI ========
    base = df_ct.copy()
    base["DataInizio"] = _to_dt(base["DataInizio"])
    base["DataFine"] = _to_dt(base["DataFine"])
    base["Stato"] = base["Stato"].astype(str).str.lower().fillna("")
    base["Durata"] = pd.to_numeric(base["Durata"], errors="coerce")
    base["TotRataNum"] = base["TotRata"].apply(_to_float_eur)

    # Join TMK dal dataframe clienti (per filtro e grafici per TMK)
    cli_tmk = df_cli[["ClienteID", "RagioneSociale", "TMK"]].copy()
    base = base.merge(cli_tmk, on="ClienteID", how="left", suffixes=("", "_cli"))
    base["TMK"] = base["TMK"].fillna("")

    today = pd.Timestamp.now().normalize()
    start_12m = (today - pd.DateOffset(months=12)).replace(day=1)

    # ======== FILTRI ========
    f1, f2, f3, f4 = st.columns([1.4, 1.2, 1.2, 1.2])
    with f1:
        periodo = st.selectbox(
            "Periodo",
            ["Ultimi 12 mesi", "Anno Corrente", "Tutto"],
            index=0
        )
    with f2:
        stato_sel = st.multiselect(
            "Stato contratto",
            options=["aperto", "chiuso", ""],
            default=["aperto", ""]  # di default mostro attivi + vuoti
        )
    with f3:
        tmk_opts = ["Tutti"] + sorted([t for t in base["TMK"].dropna().unique().tolist() if t])
        tmk_sel = st.selectbox("TMK", options=tmk_opts, index=0)
    with f4:
        solo_con_num = st.checkbox("Solo contratti con N°", value=False)

    df = base.copy()
    # Periodo → filtro su DataInizio (quando disponibile)
    if periodo == "Ultimi 12 mesi":
        df = df[(df["DataInizio"].isna()) | (df["DataInizio"] >= start_12m)]
    elif periodo == "Anno Corrente":
        df = df[(df["DataInizio"].isna()) | (df["DataInizio"] >= pd.Timestamp(today.year, 1, 1))]

    # Stato
    if stato_sel:
        df = df[df["Stato"].isin(stato_sel)]

    # TMK
    if tmk_sel != "Tutti":
        df = df[df["TMK"] == tmk_sel]

    # Solo con numero contratto
    if solo_con_num:
        df = df[df["NumeroContratto"].astype(str).str.strip() != ""]

    # ======== KPI PRINCIPALI ========
    k1, k2, k3, k4, k5 = st.columns(5)
    tot_clienti = df_cli["ClienteID"].nunique()
    attivi = int((df["Stato"] != "chiuso").sum())
    chiusi = int((df["Stato"] == "chiuso").sum())
    ytd_start = pd.Timestamp(today.year, 1, 1)
    nuovi_ytd = int(((df["DataInizio"] >= ytd_start) & df["DataInizio"].notna()).sum())
    somma_rata = df["TotRataNum"].dropna().sum()
    somma_rata_fmt = f"{somma_rata:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")

    k1.metric("Clienti totali", f"{tot_clienti}")
    k2.metric("Contratti attivi", f"{attivi}")
    k3.metric("Contratti chiusi", f"{chiusi}")
    k4.metric("Nuovi contratti YTD", f"{nuovi_ytd}")
    k5.metric("Somma rata (filtro)", somma_rata_fmt)

    st.divider()

    # ======== KPI GEOGRAFICI ========
    g1, g2, g3 = st.columns(3)
    # Città uniche
    citta_uniche = df_cli["Citta"].dropna().astype(str).str.strip().replace("", pd.NA).dropna().nunique()
    # CAP / Paesi unici (usiamo CAP come “paese” nel tuo dataset)
    cap_unici = df_cli["CAP"].dropna().astype(str).str.strip().replace("", pd.NA).dropna().nunique()
    # TMK unici (utile per capire la distribuzione commerciale)
    tmk_unici = df_cli["TMK"].dropna().astype(str).str.strip().replace("", pd.NA).dropna().nunique()

    g1.metric("🌆 Città servite", f"{citta_uniche}")
    g2.metric("🏘️ CAP / Paesi", f"{cap_unici}")
    g3.metric("👩‍💼 TMK attivi", f"{tmk_unici}")


    st.divider()
        # ======== GRAFICI GEOGRAFICI ========
    st.markdown("### 🗺️ Distribuzione geografica clienti")

    # Normalizzazione leggera campi (evita vuoti e differenze di spazi/maiuscole)
    geo = df_cli[["ClienteID", "RagioneSociale", "Citta", "CAP", "TMK"]].copy()
    for col in ["Citta", "CAP", "TMK"]:
        geo[col] = (
            geo[col]
            .fillna("")
            .astype(str)
            .str.strip()
        )

    # Opzioni di visualizzazione
    col_cfg1, col_cfg2, col_cfg3 = st.columns([1.2, 1.2, 1.2])
    with col_cfg1:
        top_n = st.slider("Top N (città/CAP)", min_value=5, max_value=30, value=10, step=1)
    with col_cfg2:
        soglia_min = st.number_input("Soglia minima clienti", min_value=0, value=1, step=1)
    with col_cfg3:
        mostra_altre = st.checkbox("Mostra anche categoria 'Altre'", value=True)

    # -------- Città --------
    st.markdown("#### 🌆 Clienti per Città")
    città_counts = (
        geo[geo["Citta"] != ""]
        .groupby("Citta")["ClienteID"].nunique()
        .sort_values(ascending=False)
        .rename("Clienti")
        .to_frame()
    )

    if città_counts.empty:
        st.info("Nessuna città disponibile nei dati clienti.")
    else:
        # Filtri top/soglia
        città_filtrate = città_counts[città_counts["Clienti"] >= soglia_min]
        top_città = città_filtrate.head(top_n)
        if mostra_altre and len(città_filtrate) > top_n:
            altre = città_filtrate.iloc[top_n:]
            altre_sum = int(altre["Clienti"].sum())
            top_città.loc["Altre"] = altre_sum

        st.bar_chart(top_città, use_container_width=True)

        # Tabella di dettaglio con link rapido “Apri”
        with st.expander("📋 Elenco (città → clienti)"):
            for c, row in top_città.sort_values("Clienti", ascending=False).iterrows():
                col_a, col_b, col_c = st.columns([2, 1, 0.8])
                col_a.markdown(f"**{c}**")
                col_b.markdown(f"{int(row['Clienti'])} clienti")
                if c != "Altre":
                    if col_c.button("📂 Apri", key=f"open_city_{c}"):
                        # Salvo un filtro soft per riutilizzarlo nella pagina Lista Clienti
                        st.session_state["filter_city"] = c
                        st.session_state["nav_target"] = "📋 Lista Clienti"
                        st.rerun()

    st.divider()

    # -------- CAP --------
    st.markdown("#### 🏘️ Clienti per CAP")
    cap_counts = (
        geo[geo["CAP"] != ""]
        .groupby("CAP")["ClienteID"].nunique()
        .sort_values(ascending=False)
        .rename("Clienti")
        .to_frame()
    )

    if cap_counts.empty:
        st.info("Nessun CAP disponibile nei dati clienti.")
    else:
        cap_filtrati = cap_counts[cap_counts["Clienti"] >= soglia_min]
        top_cap = cap_filtrati.head(top_n)
        if mostra_altre and len(cap_filtrati) > top_n:
            altre_cap = cap_filtrati.iloc[top_n:]
            altre_cap_sum = int(altre_cap["Clienti"].sum())
            top_cap.loc["Altri CAP"] = altre_cap_sum

        st.bar_chart(top_cap, use_container_width=True)

        with st.expander("📋 Elenco (CAP → clienti)"):
            for cap, row in top_cap.sort_values("Clienti", ascending=False).iterrows():
                col_a, col_b, col_c = st.columns([2, 1, 0.8])
                col_a.markdown(f"**{cap}**")
                col_b.markdown(f"{int(row['Clienti'])} clienti")
                if cap not in ["Altri CAP"]:
                    if col_c.button("📂 Apri", key=f"open_cap_{cap}"):
                        st.session_state["filter_cap"] = cap
                        st.session_state["nav_target"] = "📋 Lista Clienti"
                        st.rerun()

    # ======== GRAFICO: nuovi contratti ultimi 12 mesi ========
    st.markdown("#### 📈 Nuovi contratti (ultimi 12 mesi)")
    df_m = base.copy()
    df_m = df_m[(df_m["DataInizio"] >= start_12m) & df_m["DataInizio"].notna()]
    serie = (
        df_m
        .assign(Mese=lambda x: x["DataInizio"].dt.to_period("M").dt.to_timestamp())
        .groupby("Mese")["NumeroContratto"].count()
        .reindex(pd.period_range(start=start_12m, end=today, freq="M").to_timestamp(), fill_value=0)
        .rename("NuoviContratti")
        .to_frame()
    )
    st.line_chart(serie, use_container_width=True)

    # ======== GRAFICO: scadenze prossimi 6 mesi ========
    st.markdown("#### 📅 Scadenze nei prossimi 6 mesi")
    entro6 = today + pd.DateOffset(months=6)
    scad = base.copy()
    scad = scad[
        (scad["DataFine"].notna()) &
        (scad["DataFine"] >= today) &
        (scad["DataFine"] <= entro6) &
        (scad["Stato"] != "chiuso")
    ]
    if scad.empty:
        st.info("Nessun contratto in scadenza nei prossimi 6 mesi.")
    else:
        serie_s = (
            scad
            .assign(Mese=lambda x: x["DataFine"].dt.to_period("M").dt.to_timestamp())
            .groupby("Mese")["NumeroContratto"].count()
            .rename("Scadenze")
            .to_frame()
        )
        st.bar_chart(serie_s, use_container_width=True)

    # ======== GRAFICO: somma rata per mese (rolling 12) ========
    st.markdown("#### 💰 Totale rata per mese (contratti filtrati)")
    df_r = df.copy()
    if df_r["DataInizio"].notna().any():
        # Costruiamo una serie per mese sul range coperto dal filtro
        start_range = (df_r["DataInizio"].min() or today).to_period("M").to_timestamp() if df_r["DataInizio"].notna().any() else start_12m
        end_range = today
        # Consideriamo la rata come “rilevata” al mese di DataInizio (indicativo)
        serie_r = (
            df_r[df_r["DataInizio"].notna()]
            .assign(Mese=lambda x: x["DataInizio"].dt.to_period("M").dt.to_timestamp())
            .groupby("Mese")["TotRataNum"].sum()
            .reindex(pd.period_range(start=start_range, end=end_range, freq="M").to_timestamp(), fill_value=0.0)
            .rename("TotRata")
            .to_frame()
        )
        st.area_chart(serie_r, use_container_width=True)
    else:
        st.info("Dati mese non sufficienti per costruire la serie della rata.")

    # ======== GRAFICO: contratti per TMK ========
    st.markdown("#### 👩‍💼 Contratti per TMK (filtrati)")
    by_tmk = df.groupby("TMK")["NumeroContratto"].count().sort_values(ascending=False)
    if by_tmk.empty:
        st.info("Nessun dato TMK per i contratti filtrati.")
    else:
        st.bar_chart(by_tmk.rename("Contratti"), use_container_width=True)

    # ======== GRAFICO: distribuzione durate ========
    st.markdown("#### ⏳ Durata contratti (mesi)")
    durata_counts = df["Durata"].dropna().astype(int, errors="ignore")
    if durata_counts.empty:
        st.info("Nessuna durata disponibile nei contratti filtrati.")
    else:
        dur_tab = durata_counts.value_counts().sort_index().rename("Contratti").to_frame()
        st.bar_chart(dur_tab, use_container_width=True)

    st.divider()

    # ======== ANOMALIE E QUALITÀ DATI ========
    st.markdown("#### 🧪 Controlli qualità dati")
    colA, colB, colC = st.columns(3)

    senza_fine = df[(df["DataInizio"].notna()) & (df["DataFine"].isna())]
    colA.metric("Senza DataFine (filtrati)", f"{len(senza_fine)}")

    fine_prima = df[(df["DataInizio"].notna()) & (df["DataFine"].notna()) & (df["DataFine"] < df["DataInizio"])]
    colB.metric("DataFine < DataInizio", f"{len(fine_prima)}")

    zero_rata = df[df["TotRataNum"].fillna(0) <= 0]
    colC.metric("TotRata nulla/assente", f"{len(zero_rata)}")

    with st.expander("📋 Apri liste anomalie"):
        t1, t2, t3 = st.tabs(["Senza DataFine", "Fine < Inizio", "Rata nulla/assente"])
        with t1:
            st.dataframe(senza_fine[["ClienteID", "RagioneSociale", "NumeroContratto", "DataInizio", "DataFine", "TMK"]], use_container_width=True)
        with t2:
            st.dataframe(fine_prima[["ClienteID", "RagioneSociale", "NumeroContratto", "DataInizio", "DataFine", "TMK"]], use_container_width=True)
        with t3:
            st.dataframe(zero_rata[["ClienteID", "RagioneSociale", "NumeroContratto", "TotRata", "TMK"]], use_container_width=True)

    st.caption("Suggerimento: usa i filtri in alto per affinare i grafici e i controlli.")

# =====================================
# PAGINA RECALL E VISITE (aggiornata e coerente)
# =====================================
def page_richiami_visite(df_cli: pd.DataFrame, df_ct: pd.DataFrame, role: str):
    st.image(LOGO_URL, width=120)
    st.markdown("<h2>📅 Gestione Recall e Visite</h2>", unsafe_allow_html=True)
    st.divider()

    col1, col2 = st.columns(2)
    filtro_nome = col1.text_input("🔍 Cerca per nome cliente")
    filtro_citta = col2.text_input("🏙️ Cerca per città")

    df = df_cli.copy()
    if filtro_nome:
        df = df[df["RagioneSociale"].str.contains(filtro_nome, case=False, na=False)]
    if filtro_citta:
        df = df[df["Citta"].str.contains(filtro_citta, case=False, na=False)]
    if df.empty:
        st.warning("❌ Nessun cliente trovato.")
        return

    oggi = pd.Timestamp.now().normalize()
    for c in ["UltimoRecall", "ProssimoRecall", "UltimaVisita", "ProssimaVisita"]:
        df[c] = pd.to_datetime(df[c], errors="coerce", dayfirst=True)

    # === Imminenti (entro 30 giorni) ===
    st.markdown("### 🔔 Recall e Visite imminenti (entro 30 giorni)")
    imminenti = df[
        (df["ProssimoRecall"].between(oggi, oggi + pd.DateOffset(days=30))) |
        (df["ProssimaVisita"].between(oggi, oggi + pd.DateOffset(days=30)))
    ]

    if imminenti.empty:
        st.success("✅ Nessun richiamo o visita imminente.")
    else:
        for i, r in imminenti.iterrows():
            c1, c2, c3, c4 = st.columns([2, 1, 1, 0.7])
            c1.markdown(f"**{r['RagioneSociale']}**")
            c2.markdown(fmt_date(r["ProssimoRecall"]))
            c3.markdown(fmt_date(r["ProssimaVisita"]))
            if c4.button("📂 Apri", key=f"imm_{i}", use_container_width=True):
                st.session_state.update({
                    "selected_cliente": r["ClienteID"],
                    "nav_target": "Clienti",
                    "_go_clienti_now": True
                })
                st.rerun()

    st.divider()

    # === Recall e visite in ritardo ===
    st.markdown("### ⚠️ Recall e Visite scaduti")
    recall_vecchi = df[
        df["UltimoRecall"].notna() & (df["UltimoRecall"] < oggi - pd.DateOffset(months=3))
    ]
    visite_vecchie = df[
        df["UltimaVisita"].notna() & (df["UltimaVisita"] < oggi - pd.DateOffset(months=6))
    ]

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("#### 📞 Recall > 3 mesi fa")
        if recall_vecchi.empty:
            st.info("✅ Nessun recall scaduto.")
        else:
            for i, r in recall_vecchi.iterrows():
                c1, c2, c3 = st.columns([2.5, 1.2, 0.8])
                c1.markdown(f"**{r['RagioneSociale']}**")
                c2.markdown(fmt_date(r["UltimoRecall"]))
                if c3.button("📂 Apri", key=f"rec_{i}", use_container_width=True):
                    st.session_state.update({
                        "selected_cliente": r["ClienteID"],
                        "nav_target": "Clienti",
                        "_go_clienti_now": True
                    })
                    st.rerun()

    with col2:
        st.markdown("#### 👣 Visite > 6 mesi fa")
        if visite_vecchie.empty:
            st.info("✅ Nessuna visita scaduta.")
        else:
            for i, r in visite_vecchie.iterrows():
                c1, c2, c3 = st.columns([2.5, 1.2, 0.8])
                c1.markdown(f"**{r['RagioneSociale']}**")
                c2.markdown(fmt_date(r["UltimaVisita"]))
                if c3.button("📂 Apri", key=f"vis_{i}", use_container_width=True):
                    st.session_state.update({
                        "selected_cliente": r["ClienteID"],
                        "nav_target": "Clienti",
                        "_go_clienti_now": True
                    })
                    st.rerun()

    st.divider()
    st.markdown("### 🧾 Storico Recall e Visite")
    tabella = df[["RagioneSociale", "UltimoRecall", "ProssimoRecall", "UltimaVisita", "ProssimaVisita"]].copy()
    for c in ["UltimoRecall", "ProssimoRecall", "UltimaVisita", "ProssimaVisita"]:
        tabella[c] = tabella[c].apply(fmt_date)
    st.dataframe(tabella, use_container_width=True, hide_index=True)



# =====================================
# 📇 PAGINA LISTA COMPLETA CLIENTI E SCADENZE (CON FILTRO TMK)
# =====================================
def page_lista_clienti(df_cli: pd.DataFrame, df_ct: pd.DataFrame, role: str):
    st.title("📋 Lista Completa Clienti e Scadenze Contratti")
    oggi = pd.Timestamp.now().normalize()

    # === Prepara i dati contratti ===
    df_ct = df_ct.copy()
    df_ct["DataFine"] = pd.to_datetime(df_ct["DataFine"], errors="coerce", dayfirst=True)
    df_ct["Stato"] = df_ct["Stato"].astype(str).str.lower().fillna("")
    attivi = df_ct[df_ct["Stato"] != "chiuso"]

    # === Calcola la prima scadenza per ogni cliente ===
    prime_scadenze = (
        attivi.groupby("ClienteID")["DataFine"]
        .min()
        .reset_index()
        .rename(columns={"DataFine": "PrimaScadenza"})
    )

    merged = df_cli.merge(prime_scadenze, on="ClienteID", how="left")
    merged["GiorniMancanti"] = (merged["PrimaScadenza"] - oggi).dt.days

    # === Badge colorati per scadenza ===
    def badge_scadenza(row):
        if pd.isna(row["PrimaScadenza"]):
            return "<span style='color:#999;'>⚪ Nessuna</span>"
        giorni = row["GiorniMancanti"]
        data_fmt = fmt_date(row["PrimaScadenza"])
        if giorni < 0:
            return f"<span style='color:#757575;font-weight:600;'>⚫ Scaduto ({data_fmt})</span>"
        elif giorni <= 30:
            return f"<span style='color:#d32f2f;font-weight:600;'>🔴 {data_fmt}</span>"
        elif giorni <= 90:
            return f"<span style='color:#f9a825;font-weight:600;'>🟡 {data_fmt}</span>"
        else:
            return f"<span style='color:#388e3c;font-weight:600;'>🟢 {data_fmt}</span>"

    merged["ScadenzaBadge"] = merged.apply(badge_scadenza, axis=1)

    # === FILTRI PRINCIPALI ===
    st.markdown("### 🔍 Filtri")
    col1, col2, col3, col4, col5 = st.columns([1.5, 1.5, 1.5, 1.5, 1.5])

    filtro_nome = col1.text_input("Cerca per nome cliente")
    filtro_citta = col2.text_input("Cerca per città")
    tmk_options = ["Tutti", "Giulia", "Antonella", "Annalisa", "Laura"]
    filtro_tmk = col3.selectbox("Filtra per TMK", tmk_options, index=0)
    data_da = col4.date_input("Da data scadenza:", value=None, format="DD/MM/YYYY")
    data_a = col5.date_input("A data scadenza:", value=None, format="DD/MM/YYYY")

    # === Applica filtri ===
    if filtro_nome:
        merged = merged[merged["RagioneSociale"].str.contains(filtro_nome, case=False, na=False)]
    if filtro_citta:
        merged = merged[merged["Citta"].str.contains(filtro_citta, case=False, na=False)]
    if filtro_tmk != "Tutti":
        merged = merged[merged["TMK"] == filtro_tmk]
    if data_da:
        merged = merged[merged["PrimaScadenza"] >= pd.Timestamp(data_da)]
    if data_a:
        merged = merged[merged["PrimaScadenza"] <= pd.Timestamp(data_a)]

    # === RIEPILOGO NUMERICO ===
    total_clienti = len(merged)
    entro_30 = (merged["GiorniMancanti"] <= 30).sum()
    entro_90 = ((merged["GiorniMancanti"] > 30) & (merged["GiorniMancanti"] <= 90)).sum()
    oltre_90 = (merged["GiorniMancanti"] > 90).sum()
    scaduti = (merged["GiorniMancanti"] < 0).sum()
    senza_scadenza = merged["PrimaScadenza"].isna().sum()

    st.markdown(f"""
    **Totale Clienti:** {total_clienti}  
    ⚫ **Scaduti:** {scaduti}  
    🔴 **Entro 30 giorni:** {entro_30}  
    🟡 **Entro 90 giorni:** {entro_90}  
    🟢 **Oltre 90 giorni:** {oltre_90}  
    ⚪ **Senza scadenza:** {senza_scadenza}
    """)

    # === ORDINAMENTO ===
    st.markdown("### ↕️ Ordinamento elenco")
    ord_col1, ord_col2 = st.columns(2)
    sort_mode = ord_col1.radio(
        "Ordina per:",
        ["Nome Cliente (A → Z)", "Nome Cliente (Z → A)", "Data Scadenza (più vicina)", "Data Scadenza (più lontana)"],
        horizontal=True,
        key="sort_lista_clienti"
    )

    if sort_mode == "Nome Cliente (A → Z)":
        merged = merged.sort_values("RagioneSociale", ascending=True)
    elif sort_mode == "Nome Cliente (Z → A)":
        merged = merged.sort_values("RagioneSociale", ascending=False)
    elif sort_mode == "Data Scadenza (più vicina)":
        merged = merged.sort_values("PrimaScadenza", ascending=True, na_position="last")
    elif sort_mode == "Data Scadenza (più lontana)":
        merged = merged.sort_values("PrimaScadenza", ascending=False, na_position="last")

    # === VISUALIZZAZIONE ===
    st.divider()
    st.markdown("### 📇 Elenco Clienti e Scadenze")

    if merged.empty:
        st.warning("❌ Nessun cliente trovato con i criteri selezionati.")
        return

    for i, r in merged.iterrows():
        c1, c2, c3, c4, c5 = st.columns([2, 1.5, 1.2, 1.2, 0.7])
        with c1:
            st.markdown(f"**{r['RagioneSociale']}**")
        with c2:
            st.markdown(r.get("Citta", "") or "—")
        with c3:
            st.markdown(r["ScadenzaBadge"], unsafe_allow_html=True)
        with c4:
            tmk = r.get("TMK", "")
            if tmk:
                st.markdown(f"<span style='background:#e3f2fd;color:#0d47a1;padding:3px 8px;border-radius:8px;font-weight:600;'>{tmk}</span>", unsafe_allow_html=True)
            else:
                st.markdown("—")
        with c5:
            if st.button("📂 Apri", key=f"apri_cli_{i}", use_container_width=True):
                st.session_state.update({
                    "selected_cliente": str(r["ClienteID"]),
                    "nav_target": "Clienti",
                    "_go_clienti_now": True,
                    "_force_scroll_top": True
                })
                st.rerun()

    st.caption(f"📋 Totale clienti mostrati: **{len(merged)}**")
# =====================================
# FIX DATE: ESEGUILO UNA SOLA VOLTA
# =====================================
def fix_dates_once(df_cli: pd.DataFrame, df_ct: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Corregge le date solo una volta per sessione.
    NON usa variabili globali, evita NameError.
    Ritorna SEMPRE (df_cli, df_ct) aggiornati.
    """
    if st.session_state.get("_date_fix_done", False):
        return df_cli, df_ct

    try:
        # 🔹 Clienti
        if not df_cli.empty:
            for c in ["UltimoRecall", "ProssimoRecall", "UltimaVisita", "ProssimaVisita"]:
                if c in df_cli.columns:
                    df_cli[c] = fix_inverted_dates(df_cli[c], col_name=c)

        # 🔹 Contratti
        if not df_ct.empty:
            for c in ["DataInizio", "DataFine"]:
                if c in df_ct.columns:
                    df_ct[c] = fix_inverted_dates(df_ct[c], col_name=c)

        # 🔹 Salva una sola volta
        df_cli.to_csv(CLIENTI_CSV, index=False, encoding="utf-8-sig")
        df_ct.to_csv(CONTRATTI_CSV, index=False, encoding="utf-8-sig")

        st.toast("🔄 Date corrette e salvate nei CSV.", icon="✅")
        st.session_state["_date_fix_done"] = True
    except Exception as e:
        st.warning(f"⚠️ Correzione automatica date non completata: {e}")

    return df_cli, df_ct

# =====================================
# MAIN APP — versione 2025 GitHub + Streamlit Cloud (multi-proprietario)
# =====================================
def main():
    st.write("✅ Avvio CRM SHT — Buon Lavoro")

    # --- LOGIN (mostra schermata se non autenticato) ---
    if not st.session_state.get("logged_in", False):
        do_login_fullscreen()
        st.stop()

    user = st.session_state.get("user", "")
    role = st.session_state.get("role", "")

    if not user:
        st.warning("⚠️ Nessun utente loggato — ricarica la pagina.")
        st.stop()

    # --- Ruolo e diritti di scrittura ---
    if user == "fabio":
        ruolo_scrittura = "full"
    elif user in ["emanuela", "claudia"]:
        ruolo_scrittura = "full"
    elif user in ["giulia", "antonella", "gabriele", "laura", "annalisa"]:
        ruolo_scrittura = "limitato"
    else:
        ruolo_scrittura = "limitato"

    # --- Selettore visibilità ---
    if user in ["fabio", "giulia", "antonella", "emanuela", "claudia"]:
        visibilita_opzioni = ["Fabio", "Gabriele", "Tutti"]
        visibilita_scelta = st.sidebar.radio(
            "📂 Visualizza clienti di:",
            visibilita_opzioni,
            index=0
        )
    else:
        visibilita_scelta = "Fabio"

    # --- Caricamento dati base (Fabio) ---
    try:
        # 🔹 tenta di leggere dal database MySQL
        df_cli_main = load_table("clienti")
        df_ct_main = load_table("contratti_clienti")
        st.success("✅ Dati caricati da MySQL")
    except Exception as e:
        st.error(f"⚠️ Errore connessione MySQL: {e}")
        st.warning("Uso temporaneamente i CSV locali (solo in memoria).")
        df_cli_main = load_clienti()
        df_ct_main = load_contratti()

    # --- Caricamento dati Gabriele ---
    try:
        if GABRIELE_CLIENTI.exists():
            for sep_try in [";", ",", "|", "\t"]:
                try:
                    df_cli_gab = pd.read_csv(
                        GABRIELE_CLIENTI,
                        dtype=str,
                        sep=sep_try,
                        encoding="utf-8-sig",
                        on_bad_lines="skip",
                        engine="python"
                    ).fillna("")
                    if len(df_cli_gab.columns) > 3:
                        break
                except Exception:
                    continue
        else:
            df_cli_gab = pd.DataFrame(columns=CLIENTI_COLS)

        if GABRIELE_CONTRATTI.exists():
            for sep_try in [";", ",", "|", "\t"]:
                try:
                    df_ct_gab = pd.read_csv(
                        GABRIELE_CONTRATTI,
                        dtype=str,
                        sep=sep_try,
                        encoding="utf-8-sig",
                        on_bad_lines="skip",
                        engine="python"
                    ).fillna("")
                    if len(df_ct_gab.columns) > 3:
                        break
                except Exception:
                    continue
        else:
            df_ct_gab = pd.DataFrame(columns=CONTRATTI_COLS)

        # 🔹 Correzione colonne mancanti (solo in memoria)
        df_cli_gab = ensure_columns(df_cli_gab, CLIENTI_COLS)
        df_ct_gab = ensure_columns(df_ct_gab, CONTRATTI_COLS)

    except Exception as e:
        st.warning(f"⚠️ Impossibile caricare i dati di Gabriele: {e}")
        df_cli_gab = pd.DataFrame(columns=CLIENTI_COLS)
        df_ct_gab = pd.DataFrame(columns=CONTRATTI_COLS)


    # --- Applica filtro visibilità ---
    if visibilita_scelta == "Fabio":
        df_cli, df_ct = df_cli_main, df_ct_main
    elif visibilita_scelta == "Gabriele":
        df_cli, df_ct = df_cli_gab, df_ct_gab
    else:
        df_cli = pd.concat([df_cli_main, df_cli_gab], ignore_index=True)
        df_ct = pd.concat([df_ct_main, df_ct_gab], ignore_index=True)

    # --- Correzione date automatica una sola volta ---
    df_cli, df_ct = fix_dates_once(df_cli, df_ct)

    # --- Sidebar info ---
    st.sidebar.success(f"👤 {user} — Ruolo: {role}")
    st.sidebar.info(f"📂 Vista: {visibilita_scelta}")

    # --- Salva contesto in sessione ---
    st.session_state["ruolo_scrittura"] = ruolo_scrittura
    st.session_state["visibilita"] = visibilita_scelta

    # --- Pagine principali ---
    PAGES = {
        "Dashboard": page_dashboard,
        "📈 Dashboard Grafici": page_dashboard_grafici,
        "Clienti": page_clienti,
        "Contratti": page_contratti,
        "✏️ Modifica Contratto": page_modifica_contratto,
        "📅 Recall e Visite": page_richiami_visite,
        "📋 Lista Clienti": page_lista_clienti,
    }

    # --- Menu laterale ---
    page = st.sidebar.radio("📂 Menu principale", list(PAGES.keys()), index=0)

    # --- Navigazione automatica (dai pulsanti interni) ---
    if "nav_target" in st.session_state:
        target = st.session_state.pop("nav_target")  # ✅ pop rimuove la chiave dopo l’uso
        if target in PAGES:
            page = target


    # --- Esecuzione pagina selezionata ---
    if page in PAGES:
        st.session_state["utente_loggato"] = user
        PAGES[page](df_cli, df_ct, ruolo_scrittura)


# =====================================
# AVVIO APP
# =====================================
if __name__ == "__main__":
    main()
