# ==========================================
# import_xlsm_to_csv.py
# Ricostruisce storage/clienti.csv dal file GESTIONE_CLIENTI.xlsm
# Estrae anche le NOTE CLIENTI dai singoli fogli, ovunque siano nel file
# ==========================================

import pandas as pd
import openpyxl
from pathlib import Path

# === Percorsi ===
BASE_DIR = Path(__file__).resolve().parent
SRC_FILE = BASE_DIR / "GESTIONE_CLIENTI.xlsm"
OUT_DIR = BASE_DIR / "storage"
OUT_DIR.mkdir(exist_ok=True)
OUT_CSV = OUT_DIR / "clienti.csv"

print("📘 Caricamento file Excel:", SRC_FILE)

# === Carica workbook Excel ===
wb = openpyxl.load_workbook(SRC_FILE, data_only=True)
sheets = wb.sheetnames
print(f"🔍 Trovati {len(sheets)} fogli nel file...")

records = []

# === Loop su ogni foglio ===
for sheet_name in sheets:
    if sheet_name.strip().lower() in ["indice", "statistiche", "cap_lista", "nuovocontratto", "log_aggiornamenti"]:
        continue

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    record = {
        "ClienteID": "",
        "RagioneSociale": sheet_name.strip(),
        "PersonaRiferimento": "",
        "Indirizzo": "",
        "Citta": "",
        "CAP": "",
        "Telefono": "",
        "Cell": "",
        "Email": "",
        "PartitaIVA": "",
        "IBAN": "",
        "SDI": "",
        "UltimoRecall": "",
        "ProssimoRecall": "",
        "UltimaVisita": "",
        "ProssimaVisita": "",
        "NoteCliente": "",
    }

    # === Ricerca campi anagrafici base ===
    for row in rows:
        r = [str(c).strip() if c else "" for c in row]
        line = " ".join(r).lower()

        if "nome cliente" in line and len(r) > 1:
            record["RagioneSociale"] = r[1]
        elif "indirizzo" in line and len(r) > 1:
            record["Indirizzo"] = r[1]
        elif "citt" in line and len(r) > 1:
            record["Citta"] = r[1]
        elif "cap" in line and len(r) > 1:
            record["CAP"] = r[1]
        elif "telefono" in line and len(r) > 1:
            record["Telefono"] = r[1]
        elif "mail" in line and len(r) > 1:
            record["Email"] = r[1]
        elif "rif" in line and len(r) > 1 and not record["PersonaRiferimento"]:
            record["PersonaRiferimento"] = r[1]
        elif "partita iva" in line and len(r) > 1:
            record["PartitaIVA"] = r[1]
        elif "sdi" in line and len(r) > 1:
            record["SDI"] = r[1]
        elif "ultimo recall" in line and len(r) > 1:
            record["UltimoRecall"] = r[1]
        elif "ultima visita" in line and len(r) > 1:
            record["UltimaVisita"] = r[1]

    # === Trova e importa NOTE CLIENTI (versione flessibile) ===
    for i, row in enumerate(rows):
        line = " ".join(str(c).strip().lower() for c in row if c)
        if "note" in line and "client" in line:  # intercetta 'NOTE CLIENTI :'
            notes = []
            # Legge tutte le righe successive finché trova testo
            for nxt in rows[i + 1:]:
                txt = " ".join(str(x).strip() for x in nxt if x).strip()
                if not txt:
                    break  # riga vuota → fine note
                # se la riga è un nuovo titolo o sezione → stop
                if "contratti" in txt.lower() or ("cliente" in txt.lower() and "note" not in txt.lower()):
                    break
                notes.append(txt)
            record["NoteCliente"] = " ".join(notes)
            break

    records.append(record)

# === Assegna ID progressivo ===
for i, r in enumerate(records, start=1):
    r["ClienteID"] = str(i)

# === Esporta CSV ===
df = pd.DataFrame(records)
df.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")

# === Riepilogo finale ===
tot = len(df)
con_note = df["NoteCliente"].astype(str).str.strip().replace("nan", "").replace("None", "").ne("").sum()
senza_note = tot - con_note

print("\n✅ File esportato con successo!")
print(f"📁 Percorso: {OUT_CSV}")
print(f"👥 Clienti totali: {tot}")
print(f"📝 Con note: {con_note}")
print(f"⚪ Senza note: {senza_note}")
print("✅ Importazione completata con successo!\n")
