# =====================================
# utils/exports.py — funzioni di esportazione (Excel + PDF)
# =====================================
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from utils.formatting import fmt_date, safe_text
from utils.pdf_builder import SHTPDF


# =====================================
# 📘 ESPORTAZIONE IN EXCEL
# =====================================
def export_excel_contratti(df_ct, sel_id, rag_soc):
    """Esporta i contratti di un cliente in formato Excel A4 orizzontale con stile professionale."""
    disp = df_ct[df_ct["ClienteID"].astype(str) == str(sel_id)].copy()
    disp["DataInizio"] = disp["DataInizio"].apply(fmt_date)
    disp["DataFine"] = disp["DataFine"].apply(fmt_date)

    # ✅ colonne principali
    headers = [
        "NumeroContratto", "DataInizio", "DataFine", "Durata",
        "DescrizioneProdotto", "NOL_FIN", "NOL_INT", "TotRata",
        "CopieBN", "EccBN", "CopieCol", "EccCol", "Stato"
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = f"Contratti {rag_soc}"

    # Titolo principale
    ws.merge_cells("A1:M1")
    title = ws["A1"]
    title.value = f"Contratti Cliente: {rag_soc}"
    title.font = Font(size=14, bold=True, color="2563EB")
    title.alignment = Alignment(horizontal="center", vertical="center")

    # Riga di intestazione (blu SHT)
    ws.append(headers)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2563EB")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Applica stile all'intestazione
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i)
        c.font = head_font
        c.fill = head_fill
        c.alignment = center
        c.border = thin

    # Righe di dati
    for _, row in disp.iterrows():
        ws.append([str(row.get(h, "")) for h in headers])
        stato = str(row.get("Stato", "")).lower()
        r_idx = ws.max_row

        for j in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=j)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin
            # Riga colorata se "chiuso"
            if stato == "chiuso":
                cell.fill = PatternFill("solid", fgColor="FFCDD2")

    # ✅ Larghezze perfette per A4 orizzontale
    col_widths = [18, 14, 14, 10, 35, 14, 14, 14, 12, 12, 12, 12, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Congela riga intestazione
    ws.freeze_panes = "A3"

    # Imposta layout stampa A4 orizzontale
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.6
    ws.page_margins.bottom = 0.6
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # Esporta come bytes
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# =====================================
# EXPORT CONTRATTI → PDF (UTF-8 SAFE)
# =====================================
from fpdf import FPDF
import pandas as pd
from io import BytesIO

def export_pdf_contratti(df_ct: pd.DataFrame, sel_id: str, rag_soc: str):
    """Genera PDF dei contratti cliente — compatibile UTF-8 senza errori di codifica"""
    df = df_ct[df_ct["ClienteID"].astype(str) == str(sel_id)]
    if df.empty:
        return None

    # --- Imposta il PDF ---
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, f"Contratti Cliente: {rag_soc}", ln=True, align="C")

    pdf.ln(8)
    pdf.set_font("Helvetica", "B", 10)
    headers = ["Numero", "Data Inizio", "Data Fine", "Durata", "Descrizione", "Tot Rata", "Stato"]
    col_widths = [25, 30, 30, 20, 120, 25, 25]

    # --- Intestazione tabella ---
    for h, w in zip(headers, col_widths):
        pdf.cell(w, 8, h, border=1, align="C")
    pdf.ln()

    # --- Righe contratti ---
    pdf.set_font("Helvetica", "", 9)
    for _, row in df.iterrows():
        descr = str(row.get("DescrizioneProdotto", "")).replace("\n", " ")
        if len(descr) > 90:
            descr = descr[:90] + "…"
        valori = [
            str(row.get("NumeroContratto", "")),
            str(row.get("DataInizio", "")),
            str(row.get("DataFine", "")),
            str(row.get("Durata", "")),
            descr,
            str(row.get("TotRata", "")),
            str(row.get("Stato", ""))
        ]
        for val, w in zip(valori, col_widths):
            pdf.cell(w, 7, val.encode("latin-1", "replace").decode("latin-1"), border=1)
        pdf.ln()

    # --- Output buffer UTF-8 safe ---
    buffer = BytesIO()
    pdf_bytes = pdf.output(dest="S").encode("latin-1", "replace")
    buffer.write(pdf_bytes)
    buffer.seek(0)
    return buffer
